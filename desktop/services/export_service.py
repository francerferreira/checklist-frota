from __future__ import annotations

import csv
import io
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.legends import Legend
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.shapes import Drawing, Rect, String
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Image, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from runtime_paths import exports_path


BRAND_BLUE = "#0B1220"
PRIMARY_BLUE = "#2563EB"
LIGHT_BG = "#F8FAFC"
MUTED = "#64748B"
SEVERITY_RED = "#B91C1C"
SEVERITY_RED_BG = "#FEE2E2"
SEVERITY_YELLOW = "#B45309"
SEVERITY_YELLOW_BG = "#FEF3C7"
SEVERITY_GREEN = "#166534"
SEVERITY_GREEN_BG = "#DCFCE7"


def _timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _default_export_path(prefix: str, suffix: str) -> Path:
    export_dir = exports_path()
    return export_dir / f"{prefix}_{_timestamp()}.{suffix}"


def make_default_export_path(prefix: str, suffix: str) -> str:
    return str(_default_export_path(prefix, suffix))


def _origin_photo_path(record: dict | None) -> str | None:
    payload = record or {}
    return payload.get("foto_origem") or payload.get("foto_antes")


def _resolution_photo_path(record: dict | None) -> str | None:
    payload = record or {}
    return payload.get("foto_resolucao") or payload.get("foto_depois")


def export_rows_to_csv(
    columns: list[tuple[str, str]],
    rows: list[dict],
    output_path: str | Path,
) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle, delimiter=";")
        writer.writerow([label for label, _ in columns])
        for row in rows:
            writer.writerow([_stringify(row.get(key)) for _, key in columns])
    return path


def export_rows_to_xlsx(
    title: str,
    columns: list[tuple[str, str]],
    rows: list[dict],
    output_path: str | Path,
) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Relatório"

    sheet["A1"] = title
    sheet["A1"].font = Font(bold=True, size=15, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="0B1220")
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))

    for column_index, (label, _) in enumerate(columns, start=1):
        cell = sheet.cell(row=3, column=column_index, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for row_index, row in enumerate(rows, start=4):
        for column_index, (_, key) in enumerate(columns, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=_stringify(row.get(key)))
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for column_index, (label, key) in enumerate(columns, start=1):
        values = [_stringify(row.get(key)) for row in rows]
        max_size = max([len(label)] + [len(value) for value in values] + [12])
        sheet.column_dimensions[get_column_letter(column_index)].width = min(max_size + 2, 36)

    workbook.save(path)
    return path


def export_rows_to_pdf(
    title: str,
    subtitle: str,
    columns: list[tuple[str, str]],
    rows: list[dict],
    output_path: str | Path,
    *,
    logo_path: str | Path | None = None,
    generated_by: str = "",
    period_label: str | None = None,
) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    doc = SimpleDocTemplate(
        str(path),
        pagesize=landscape(A4),
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=40 * mm,
        bottomMargin=12 * mm,
    )

    styles = _styles()
    story = _build_cover_page(title, subtitle, generated_by, logo_path, styles, landscape_mode=True)
    story.append(PageBreak())
    story.extend(
        _build_summary_cards(
            [
                ("Total de linhas", str(len(rows))),
                ("Colunas", str(len(columns))),
                ("Período", period_label or f"Base consolidada até {datetime.now().strftime('%d/%m/%Y')}"),
                ("Semáforo executivo", _overall_priority_label(columns, rows)),
            ],
            styles,
        )
    )
    story.append(Spacer(1, 6))
    story.append(Paragraph("Visão analítica consolidada", styles["section"]))
    story.append(Spacer(1, 4))

    chart_section = _build_chart_section(columns, rows, styles)
    if chart_section:
        story.extend(chart_section)
        story.append(Spacer(1, 8))

    ranking_section = _build_top_five_section(columns, rows, styles)
    if ranking_section:
        story.extend(ranking_section)
        story.append(Spacer(1, 8))

    table_data = [[Paragraph(label, styles["table_header"]) for label, _ in columns]]
    for row in rows:
        table_data.append(
            [Paragraph(_safe_paragraph_text(_stringify(row.get(key))), styles["table_cell"]) for _, key in columns]
        )

    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(PRIMARY_BLUE)),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("LINEBELOW", (0, 0), (-1, -1), 0.35, colors.HexColor("#E2E8F0")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor(LIGHT_BG)]),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )

    story.append(table)
    story.append(Spacer(1, 10))
    story.extend(_build_executive_conclusion(columns, rows, styles, period_label))
    story.append(Spacer(1, 10))
    story.extend(_build_signature_block(generated_by, styles))
    page_frame = _page_frame_callback(title, subtitle, generated_by, logo_path)
    doc.build(
        story,
        onFirstPage=page_frame,
        onLaterPages=page_frame,
    )
    return path


def export_non_conformity_pdf(
    item: dict,
    *,
    output_path: str | Path,
    logo_path: str | Path | None = None,
    generated_by: str = "",
    before_image: bytes | None = None,
    after_image: bytes | None = None,
) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    doc = SimpleDocTemplate(
        str(path),
        pagesize=A4,
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=40 * mm,
        bottomMargin=12 * mm,
    )
    styles = _styles()
    story = _build_cover_page(
        "Relatório de Não Conformidade",
        f"{item['veiculo']['frota']} - {item['item_nome']}",
        generated_by,
        logo_path,
        styles,
        landscape_mode=False,
    )
    story.append(PageBreak())

    status_text = "Resolvida" if item.get("resolvido") else "Aberta"
    story.extend(
        _build_summary_cards(
            [
                ("Status", status_text),
                ("Veículo", item["veiculo"].get("frota") or "-"),
                ("Motorista", item["usuario"].get("nome") or "-"),
            ],
            styles,
        )
    )
    story.append(Spacer(1, 8))
    story.append(Paragraph("Resumo executivo da ocorrência", styles["section"]))
    story.append(Spacer(1, 4))
    info_rows = [
        ["Veículo", item["veiculo"].get("frota") or "-"],
        ["Tipo", item["veiculo"].get("tipo", "-").title()],
        ["Placa", item["veiculo"].get("placa") or "-"],
        ["Modelo", item["veiculo"].get("modelo") or "-"],
        ["Item", item.get("item_nome") or "-"],
        ["Status", status_text],
        ["Motorista", item["usuario"].get("nome") or "-"],
        ["Data de abertura", _format_datetime(item.get("created_at"))],
        ["Data de resolução", _format_datetime(item.get("data_resolucao"))],
        ["Código da peça", item.get("codigo_peca") or "-"],
        ["Descrição da peça", item.get("descricao_peca") or "-"],
        ["Observação", item.get("observacao") or "-"],
    ]
    story.append(_key_value_table(info_rows, styles))
    story.append(Spacer(1, 8))

    image_block = []
    if before_image:
        image_block.append(_reportlab_image(before_image, 86 * mm, 76 * mm))
    else:
        image_block.append(Paragraph("Sem foto de origem", styles["muted_box"]))
    if after_image:
        image_block.append(_reportlab_image(after_image, 86 * mm, 76 * mm))
    else:
        image_block.append(Paragraph("Sem foto de resolução", styles["muted_box"]))

    image_table = Table(
        [
            [
                Paragraph("Foto de origem", styles["section"]),
                Paragraph("Foto de resolução", styles["section"]),
            ],
            image_block,
        ],
        colWidths=[89 * mm, 89 * mm],
    )
    image_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.white),
                ("LINEBELOW", (0, 0), (-1, 0), 0.35, colors.HexColor("#CBD5E1")),
                ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E2E8F0")),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    story.append(image_table)
    story.append(Spacer(1, 12))
    story.extend(_build_signature_block(generated_by, styles))
    page_frame = _page_frame_callback(
        "Relatório de Não Conformidade",
        f"{item['veiculo']['frota']} - {item['item_nome']}",
        generated_by,
        logo_path,
    )
    doc.build(
        story,
        onFirstPage=page_frame,
        onLaterPages=page_frame,
    )
    return path


def export_vehicle_detail_pdf(
    vehicle: dict,
    occurrences: list[dict],
    *,
    output_path: str | Path,
    logo_path: str | Path | None = None,
    generated_by: str = "",
    vehicle_image: bytes | None = None,
    operational_history: list[dict] | None = None,
    occurrence_images: dict[int, dict[str, bytes | None]] | None = None,
) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    doc = SimpleDocTemplate(
        str(path),
        pagesize=A4,
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=40 * mm,
        bottomMargin=12 * mm,
    )
    styles = _styles()
    story = _build_cover_page(
        "Ficha de Equipamento",
        f"{vehicle.get('frota', '-')} - {vehicle.get('modelo', '-')}",
        generated_by,
        logo_path,
        styles,
        landscape_mode=False,
    )
    story.append(PageBreak())
    story.extend(
        _build_summary_cards(
            [
                ("Frota", vehicle.get("frota") or "-"),
                ("Tipo", (vehicle.get("tipo") or "-").title()),
                ("Não conformidades acumuladas", str(len(occurrences))),
            ],
            styles,
        )
    )
    story.append(Spacer(1, 8))
    story.append(Paragraph("Resumo técnico do ativo", styles["section"]))
    story.append(Spacer(1, 4))

    rows = [
        ["Frota", vehicle.get("frota") or "-"],
        ["Tipo", (vehicle.get("tipo") or "-").title()],
        ["Placa", vehicle.get("placa") or "-"],
        ["Modelo", vehicle.get("modelo") or "-"],
        ["Ano", vehicle.get("ano") or "-"],
        ["Chassi", vehicle.get("chassi") or "-"],
        ["Configuração", vehicle.get("configuracao") or "-"],
        ["Atividade", vehicle.get("atividade") or "-"],
        ["Status", vehicle.get("status") or "-"],
        ["Local", vehicle.get("local") or "-"],
        ["Descrição", vehicle.get("descricao") or "-"],
        ["Total de ocorrências de não conformidade", str(len(occurrences))],
    ]
    story.append(_key_value_table(rows, styles))
    story.append(Spacer(1, 8))

    if vehicle_image:
        story.append(Paragraph("Foto do equipamento", styles["section"]))
        story.append(Spacer(1, 4))
        story.append(_reportlab_image(vehicle_image, 170 * mm, 95 * mm))
        story.append(Spacer(1, 10))

    occ_columns = [
        ("Data", "created_at"),
        ("Item", "item_nome"),
        ("Status", "status_text"),
        ("Peça", "codigo_peca"),
        ("Motorista", "motorista"),
    ]
    occ_rows = []
    for item in occurrences:
        occ_rows.append(
            {
                "created_at": _format_datetime(item.get("created_at")),
                "item_nome": item.get("item_nome") or "-",
                "status_text": "Resolvida" if item.get("resolvido") else "Aberta",
                "codigo_peca": item.get("codigo_peca") or "-",
                "motorista": item.get("usuario", {}).get("nome") or "-",
            }
        )

    story.append(Spacer(1, 4))
    story.append(Paragraph("Ocorrências registradas", styles["section"]))
    story.append(Spacer(1, 4))

    if occ_rows:
        occ_table = Table(
            [[Paragraph(label, styles["table_header"]) for label, _ in occ_columns]]
            + [
                [Paragraph(_safe_paragraph_text(_stringify(row.get(key))), styles["table_cell"]) for _, key in occ_columns]
                for row in occ_rows
            ],
            repeatRows=1,
        )
        occ_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(PRIMARY_BLUE)),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("LINEBELOW", (0, 0), (-1, -1), 0.35, colors.HexColor("#E2E8F0")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor(LIGHT_BG)]),
                    ("LEFTPADDING", (0, 0), (-1, -1), 7),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                    ("TOPPADDING", (0, 0), (-1, -1), 7),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
                ]
            )
        )
        story.append(occ_table)
    else:
        story.append(Paragraph("Nenhuma não conformidade registrada para este equipamento.", styles["muted_box"]))

    if occurrences:
        _append_occurrence_evidence_section(
            story,
            occurrences,
            occurrence_images or {},
            styles,
            section_title="Evidências das não conformidades",
            page_break_before=True,
        )

    if operational_history:
        history_columns = [
            ("Data", "date"),
            ("Origem", "origin"),
            ("Item", "item"),
            ("Status", "status"),
            ("Responsável", "owner"),
        ]
        history_rows = [
            {
                "date": _format_datetime(item.get("date")),
                "origin": item.get("origin") or "-",
                "item": item.get("item") or "-",
                "status": item.get("status") or "-",
                "owner": item.get("owner") or "-",
            }
            for item in operational_history[:80]
        ]
        story.append(Spacer(1, 10))
        story.append(Paragraph("Histórico operacional", styles["section"]))
        story.append(Spacer(1, 4))
        history_table = Table(
            [[Paragraph(label, styles["table_header"]) for label, _ in history_columns]]
            + [
                [Paragraph(_safe_paragraph_text(_stringify(row.get(key))), styles["table_cell"]) for _, key in history_columns]
                for row in history_rows
            ],
            repeatRows=1,
        )
        history_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(PRIMARY_BLUE)),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("LINEBELOW", (0, 0), (-1, -1), 0.35, colors.HexColor("#E2E8F0")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor(LIGHT_BG)]),
                    ("LEFTPADDING", (0, 0), (-1, -1), 7),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                    ("TOPPADDING", (0, 0), (-1, -1), 7),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
                ]
            )
        )
        story.append(history_table)

    story.append(Spacer(1, 12))
    story.extend(_build_signature_block(generated_by, styles))
    page_frame = _page_frame_callback(
        "Ficha de Equipamento",
        f"{vehicle.get('frota', '-')} - {vehicle.get('modelo', '-')}",
        generated_by,
        logo_path,
    )
    doc.build(
        story,
        onFirstPage=page_frame,
        onLaterPages=page_frame,
    )
    return path


def export_item_audit_pdf(
    item_name: str | None,
    occurrences: list[dict],
    *,
    output_path: str | Path,
    logo_path: str | Path | None = None,
    generated_by: str = "",
    occurrence_images: dict[int, dict[str, bytes | None]] | None = None,
    filter_context: dict[str, str] | None = None,
    resolved_mode: bool = False,
    include_resolution_details: bool = False,
    include_part_details: bool = False,
) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    doc = SimpleDocTemplate(
        str(path),
        pagesize=A4,
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=40 * mm,
        bottomMargin=12 * mm,
    )
    styles = _styles()
    vehicles = sorted({(item.get("veiculo") or {}).get("frota") or "-" for item in occurrences})
    open_total = sum(1 for item in occurrences if not item.get("resolvido"))
    resolved_total = sum(1 for item in occurrences if item.get("resolvido"))
    grouped: dict[str, dict] = {}
    for occurrence in occurrences:
        raw_name = str(occurrence.get("item_nome") or "").strip()
        normalized_name = " ".join(raw_name.upper().split()) or "NÃO INFORMADA"
        bucket = grouped.setdefault(
            normalized_name,
            {
                "item_nome": normalized_name,
                "occurrences": [],
                "vehicles": set(),
                "abertas": 0,
                "resolvidas": 0,
            },
        )
        bucket["occurrences"].append(occurrence)
        vehicle_name = (occurrence.get("veiculo") or {}).get("frota") or "-"
        bucket["vehicles"].add(vehicle_name)
        if occurrence.get("resolvido"):
            bucket["resolvidas"] += 1
        else:
            bucket["abertas"] += 1
    grouped_rows = sorted(grouped.values(), key=lambda row: (-len(row["occurrences"]), row["item_nome"]))
    scope_label = item_name or "Todas as não conformidades"
    title = "Relatório de Auditoria de Não conformidades"
    subtitle = f"{scope_label} - {len(occurrences)} ocorrências em {len(grouped_rows)} NC"

    story = _build_cover_page(title, subtitle, generated_by, logo_path, styles, landscape_mode=False)
    story.append(PageBreak())
    summary_cards = [
        ("Escopo", scope_label),
        ("Tipos de NC", str(len(grouped_rows))),
        ("Ocorrências", str(len(occurrences))),
        ("Equipamentos", str(len(vehicles))),
        ("Abertas", str(open_total)),
        ("Resolvidas", str(resolved_total)),
    ]
    if filter_context:
        for key, value in filter_context.items():
            summary_cards.append((key, value or "-"))
    story.extend(
        _build_summary_cards(summary_cards, styles)
    )
    story.append(Spacer(1, 8))
    story.append(Paragraph("Consolidado por não conformidade", styles["section"]))
    story.append(Spacer(1, 4))
    if grouped_rows:
        consolidated_columns = [
            ("Não conformidade", "item_nome"),
            ("Ocorrências", "ocorrencias"),
            ("Equipamentos", "equipamentos_total"),
            ("Abertas", "abertas"),
            ("Resolvidas", "resolvidas"),
        ]
        consolidated_table = Table(
            [[Paragraph(label, styles["table_header"]) for label, _ in consolidated_columns]]
            + [
                [
                    Paragraph(_safe_paragraph_text(_stringify(value)), styles["table_cell"])
                    for value in (
                        row["item_nome"],
                        len(row["occurrences"]),
                        len(row["vehicles"]),
                        row["abertas"],
                        row["resolvidas"],
                    )
                ]
                for row in grouped_rows
            ],
            repeatRows=1,
        )
        consolidated_table.setStyle(_standard_table_style())
        story.append(consolidated_table)
    else:
        story.append(Paragraph("Nenhuma ocorrência encontrada para o filtro atual.", styles["muted_box"]))

    for group in grouped_rows:
        story.append(PageBreak())
        group_name = group["item_nome"]
        group_occurrences = sorted(group["occurrences"], key=lambda row: row.get("created_at") or "", reverse=True)
        group_vehicles = sorted(group["vehicles"])
        story.append(Paragraph(_safe_paragraph_text(f"Não conformidade: {group_name}"), styles["section"]))
        story.append(Spacer(1, 4))
        story.extend(
            _build_summary_cards(
                [
                    ("Ocorrências", str(len(group_occurrences))),
                    ("Equipamentos", str(len(group_vehicles))),
                    ("Abertas", str(group["abertas"])),
                    ("Resolvidas", str(group["resolvidas"])),
                ],
                styles,
            )
        )
        story.append(Spacer(1, 6))
        story.append(
            Paragraph(
                _safe_paragraph_text(
                    f"Equipamentos ({len(group_vehicles)}): {', '.join(group_vehicles) if group_vehicles else '-'}"
                ),
                styles["body"],
            )
        )
        story.append(Spacer(1, 8))

        group_columns = [
            ("Data", "date_value"),
            ("Equipamento", "vehicle"),
            ("Status", "status"),
            ("Motorista", "driver"),
            ("Resolvido em", "resolved_at"),
            ("Foto origem", "before"),
            ("Foto resolução", "after"),
        ]
        group_rows = []
        for occurrence in group_occurrences:
            vehicle = occurrence.get("veiculo") or {}
            user = occurrence.get("usuario") or {}
            resolved_at = _format_datetime(occurrence.get("data_resolucao"))
            row_date = resolved_at if resolved_mode and occurrence.get("data_resolucao") else _format_datetime(occurrence.get("created_at"))
            group_rows.append(
                {
                    "date_value": row_date,
                    "vehicle": vehicle.get("frota") or "-",
                    "status": "Resolvida" if occurrence.get("resolvido") else "Aberta",
                    "driver": user.get("nome") or "-",
                    "resolved_at": resolved_at,
                    "before": "Sim" if _origin_photo_path(occurrence) else "Não",
                    "after": "Sim" if _resolution_photo_path(occurrence) else "Não",
                }
            )
        group_table = Table(
            [[Paragraph(label, styles["table_header"]) for label, _ in group_columns]]
            + [
                [Paragraph(_safe_paragraph_text(_stringify(row.get(key))), styles["table_cell"]) for _, key in group_columns]
                for row in group_rows
            ],
            repeatRows=1,
        )
        group_table.setStyle(_standard_table_style())
        story.append(group_table)

        _append_occurrence_evidence_section(
            story,
            group_occurrences,
            occurrence_images or {},
            styles,
            section_title=f"Caderno de evidências - {group_name}",
            page_break_before=True,
            include_resolution_fields=include_resolution_details,
            include_part_fields=include_part_details,
        )

    story.append(Spacer(1, 12))
    story.extend(_build_signature_block(generated_by, styles))
    page_frame = _page_frame_callback(title, subtitle, generated_by, logo_path)
    doc.build(
        story,
        onFirstPage=page_frame,
        onLaterPages=page_frame,
    )
    return path


def export_activity_pdf(
    activity: dict,
    *,
    output_path: str | Path,
    logo_path: str | Path | None = None,
    generated_by: str = "",
    item_images: dict[int, dict[str, bytes | None]] | None = None,
) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    doc = SimpleDocTemplate(
        str(path),
        pagesize=A4,
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=40 * mm,
        bottomMargin=12 * mm,
    )
    styles = _styles()
    resumo = activity.get("resumo", {})
    itens = activity.get("itens", [])
    item_images = item_images or {}

    title = activity.get("titulo") or "Atividade em massa"
    subtitle = f"{activity.get('item_nome') or '-'} • {(activity.get('tipo_equipamento') or '-').title()}"

    story = _build_cover_page(title, subtitle, generated_by, logo_path, styles, landscape_mode=False)
    story.append(PageBreak())
    story.extend(
        _build_summary_cards(
            [
                ("Equipamentos", str(resumo.get("total", len(itens)))),
                ("Instalados", str(resumo.get("instalados", 0))),
                ("Não instalados", str(resumo.get("nao_instalados", 0))),
                ("Pendentes", str(resumo.get("pendentes", 0))),
            ],
            styles,
        )
    )
    story.append(Spacer(1, 8))
    story.append(Paragraph("Resumo da atividade", styles["section"]))
    story.append(Spacer(1, 4))
    story.append(
        _key_value_table(
            [
                ["Titulo", activity.get("titulo") or "-"],
                ["Modulo / componente", activity.get("item_nome") or "-"],
                ["Tipo de equipamento", (activity.get("tipo_equipamento") or "-").title()],
                ["Status", "Finalizada" if activity.get("status") == "FINALIZADA" else "Aberta"],
                ["Código da peça", activity.get("codigo_peca") or "-"],
                ["Descrição da peça", activity.get("descricao_peca") or "-"],
                ["Fornecedor", activity.get("fornecedor_peca") or "-"],
                ["Lote", activity.get("lote_peca") or "-"],
                ["Data de abertura", _format_datetime(activity.get("created_at"))],
                ["Data de fechamento", _format_datetime(activity.get("finalized_at"))],
                ["Observação geral", activity.get("observacao") or "-"],
            ],
            styles,
        )
    )
    story.append(Spacer(1, 8))

    summary_columns = [
        ("Frota", "frota"),
        ("Placa", "placa"),
        ("Modelo", "modelo"),
        ("Material", "material"),
        ("Qtd", "quantidade_peca"),
        ("Status da atividade", "status"),
        ("Instalado em", "instalado_em"),
        ("Executado por", "executado_por"),
    ]
    summary_rows = []
    for item in itens:
        veiculo = item.get("veiculo", {})
        material = item.get("material") or {}
        summary_rows.append(
            {
                "frota": veiculo.get("frota") or "-",
                "placa": veiculo.get("placa") or "-",
                "modelo": veiculo.get("modelo") or "-",
                "material": item.get("descricao_peca") or material.get("descricao") or item.get("codigo_peca") or "-",
                "quantidade_peca": str(item.get("quantidade_peca") or activity.get("quantidade_por_equipamento") or 1),
                "status": _activity_item_status_text(item.get("status_execucao")),
                "instalado_em": _format_datetime(item.get("instalado_em")),
                "executado_por": item.get("executado_por_nome") or "-",
            }
        )

    story.append(Paragraph("Consolidado por equipamento", styles["section"]))
    story.append(Spacer(1, 4))
    summary_table = Table(
        [[Paragraph(label, styles["table_header"]) for label, _ in summary_columns]]
        + [
            [Paragraph(_safe_paragraph_text(_stringify(row.get(key))), styles["table_cell"]) for _, key in summary_columns]
            for row in summary_rows
        ],
        repeatRows=1,
    )
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(PRIMARY_BLUE)),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("LINEBELOW", (0, 0), (-1, -1), 0.35, colors.HexColor("#E2E8F0")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor(LIGHT_BG)]),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    story.append(summary_table)

    if itens:
        story.append(PageBreak())
        story.append(Paragraph("Caderno de evidências por equipamento", styles["section"]))
        story.append(Spacer(1, 6))

    for index, item in enumerate(itens):
        veiculo = item.get("veiculo", {})
        material = item.get("material") or {}
        images = item_images.get(item.get("id"), {})
        before_image = images.get("before")
        after_image = images.get("after")
        status_text = _activity_item_status_text(item.get("status_execucao"))
        stamp = _activity_status_stamp(status_text, styles)

        item_header = Table(
            [[
                Paragraph(f"{veiculo.get('frota') or '-'} • {veiculo.get('modelo') or '-'}", styles["section"]),
                stamp,
            ]],
            colWidths=[132 * mm, 48 * mm],
        )
        item_header.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#EFF6FF")),
                    ("LINELEFT", (0, 0), (-1, -1), 1.0, colors.HexColor("#BFDBFE")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 10),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                    ("TOPPADDING", (0, 0), (-1, -1), 8),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ]
            )
        )
        story.append(item_header)
        story.append(Spacer(1, 6))
        story.append(
            _key_value_table(
                [
                    ["Frota", veiculo.get("frota") or "-"],
                    ["Placa", veiculo.get("placa") or "-"],
                    ["Modelo", veiculo.get("modelo") or "-"],
                    ["Material", item.get("descricao_peca") or material.get("descricao") or item.get("codigo_peca") or "-"],
                    ["Qtd. peça", str(item.get("quantidade_peca") or activity.get("quantidade_por_equipamento") or 1)],
                    ["Código da peça", item.get("codigo_peca") or activity.get("codigo_peca") or "-"],
                    ["Descrição da peça", item.get("descricao_peca") or activity.get("descricao_peca") or "-"],
                    ["Status da atividade", status_text],
                    ["Executado em", _format_datetime(item.get("instalado_em"))],
                    ["Executado por", item.get("executado_por_nome") or "-"],
                    ["Login do executor", item.get("executado_por_login") or "-"],
                    ["Observação", item.get("observacao") or "-"],
                ],
                styles,
            )
        )
        story.append(Spacer(1, 6))

        image_block = []
        image_block.append(_reportlab_image(before_image, 86 * mm, 76 * mm) if before_image else Paragraph("Sem foto de origem", styles["muted_box"]))
        image_block.append(_reportlab_image(after_image, 86 * mm, 76 * mm) if after_image else Paragraph("Sem foto de resolução", styles["muted_box"]))
        image_table = Table(
            [
                [Paragraph("Foto de origem", styles["section"]), Paragraph("Foto de resolução", styles["section"])],
                image_block,
            ],
            colWidths=[89 * mm, 89 * mm],
        )
        image_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.white),
                    ("LINEBELOW", (0, 0), (-1, 0), 0.35, colors.HexColor("#CBD5E1")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E2E8F0")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        story.append(image_table)
        if index < len(itens) - 1:
            story.append(Spacer(1, 10))
            story.append(PageBreak())

    story.append(Spacer(1, 12))
    story.extend(_build_signature_block(generated_by, styles))
    page_frame = _page_frame_callback(title, subtitle, generated_by, logo_path)
    doc.build(
        story,
        onFirstPage=page_frame,
        onLaterPages=page_frame,
    )
    return path


def export_material_report_xlsx(
    report: dict,
    *,
    output_path: str | Path,
) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Resumo"
    summary_sheet["A1"] = "Relatório de Estoque"
    summary_sheet["A1"].font = Font(bold=True, size=15, color="FFFFFF")
    summary_sheet["A1"].fill = PatternFill("solid", fgColor="0B1220")
    summary_sheet.merge_cells("A1:B1")

    resumo = report.get("resumo", {})
    periodo = report.get("periodo", {})
    summary_rows = [
        ("Data inicial", periodo.get("data_inicial") or "-"),
        ("Data final", periodo.get("data_final") or "-"),
        ("Total de materiais", resumo.get("total_materiais", 0)),
        ("Abaixo do mínimo", resumo.get("abaixo_minimo", 0)),
        ("Saldo total", resumo.get("saldo_total", 0)),
        ("Consumo total no período", resumo.get("consumo_total_periodo", 0)),
    ]
    for row_index, (label, value) in enumerate(summary_rows, start=3):
        summary_sheet.cell(row=row_index, column=1, value=label)
        summary_sheet.cell(row=row_index, column=2, value=_stringify(value))

    def fill_sheet(sheet_name: str, columns: list[tuple[str, str]], rows: list[dict]):
        sheet = workbook.create_sheet(sheet_name)
        for column_index, (label, _) in enumerate(columns, start=1):
            cell = sheet.cell(row=1, column=column_index, value=label)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2563EB")
        for row_index, row in enumerate(rows, start=2):
            for column_index, (_, key) in enumerate(columns, start=1):
                sheet.cell(row=row_index, column=column_index, value=_stringify(row.get(key)))
        for column_index, (label, key) in enumerate(columns, start=1):
            values = [_stringify(row.get(key)) for row in rows]
            max_size = max([len(label)] + [len(value) for value in values] + [12])
            sheet.column_dimensions[get_column_letter(column_index)].width = min(max_size + 2, 36)

    fill_sheet(
        "Baixo Estoque",
        [("Referência", "referencia"), ("Descrição", "descricao"), ("Aplicação", "aplicacao_tipo"), ("Estoque", "quantidade_estoque"), ("Mínimo", "estoque_minimo"), ("Déficit", "deficit")],
        report.get("baixo_estoque", []),
    )
    fill_sheet(
        "Consumo",
        [("Referência", "referencia"), ("Descrição", "descricao"), ("Consumo", "consumo_total"), ("Último consumo", "ultimo_consumo")],
        report.get("consumo_periodo", []),
    )
    fill_sheet(
        "Ranking Top 5",
        [("Referência", "referencia"), ("Descrição", "descricao"), ("Consumo", "consumo_total"), ("Último consumo", "ultimo_consumo")],
        report.get("ranking_uso", []),
    )

    workbook.save(path)
    return path


def export_material_report_pdf(
    report: dict,
    *,
    output_path: str | Path,
    logo_path: str | Path | None = None,
    generated_by: str = "",
) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    styles = _styles()
    periodo = report.get("periodo", {})
    resumo = report.get("resumo", {})
    subtitle = (
        f"Período {periodo.get('data_inicial') or '-'} a {periodo.get('data_final') or datetime.now().strftime('%d/%m/%Y')}"
    )

    doc = SimpleDocTemplate(
        str(path),
        pagesize=A4,
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=40 * mm,
        bottomMargin=12 * mm,
    )

    story = _build_cover_page("Relatório de Estoque", subtitle, generated_by, logo_path, styles, landscape_mode=False)
    story.append(PageBreak())
    story.extend(
        _build_summary_cards(
            [
                ("Total de materiais", str(resumo.get("total_materiais", 0))),
                ("Abaixo do mínimo", str(resumo.get("abaixo_minimo", 0))),
                ("Saldo total", str(resumo.get("saldo_total", 0))),
                ("Consumo no período", str(resumo.get("consumo_total_periodo", 0))),
            ],
            styles,
        )
    )
    story.append(Spacer(1, 10))

    sections = [
        (
            "Materiais abaixo do mínimo",
            [("Referência", "referencia"), ("Descrição", "descricao"), ("Aplicação", "aplicacao_tipo"), ("Estoque", "quantidade_estoque"), ("Mínimo", "estoque_minimo"), ("Déficit", "deficit")],
            report.get("baixo_estoque", []),
        ),
        (
            "Consumo no período",
            [("Referência", "referencia"), ("Descrição", "descricao"), ("Consumo", "consumo_total"), ("Último consumo", "ultimo_consumo")],
            report.get("consumo_periodo", []),
        ),
        (
            "Ranking Top 5",
            [("Referência", "referencia"), ("Descrição", "descricao"), ("Consumo", "consumo_total"), ("Último consumo", "ultimo_consumo")],
            report.get("ranking_uso", []),
        ),
    ]

    for title, columns, rows in sections:
        story.append(Paragraph(title, styles["section"]))
        story.append(Spacer(1, 4))
        if rows:
            table = Table(
                [[Paragraph(label, styles["table_header"]) for label, _ in columns]]
                + [
                    [Paragraph(_safe_paragraph_text(_stringify(row.get(key))), styles["table_cell"]) for _, key in columns]
                    for row in rows
                ],
                repeatRows=1,
            )
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(PRIMARY_BLUE)),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("LINEBELOW", (0, 0), (-1, -1), 0.35, colors.HexColor("#E2E8F0")),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor(LIGHT_BG)]),
                        ("LEFTPADDING", (0, 0), (-1, -1), 8),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                        ("TOPPADDING", (0, 0), (-1, -1), 8),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                    ]
                )
            )
            story.append(table)
        else:
            story.append(Paragraph("Sem registros para esta seção.", styles["muted_box"]))
        story.append(Spacer(1, 10))

    story.extend(_build_signature_block(generated_by, styles))
    page_frame = _page_frame_callback("Relatório de Estoque", subtitle, generated_by, logo_path)
    doc.build(
        story,
        onFirstPage=page_frame,
        onLaterPages=page_frame,
    )
    return path


def _build_header(title: str, subtitle: str, generated_by: str, logo_path: str | Path | None, styles):
    story = []
    cover_band = Table(
        [[Paragraph("Sistema de Checklist de Frota | Relatório corporativo", styles["cover_band"])]],
        colWidths=[180 * mm],
    )
    cover_band.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor(MUTED)),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("LINEBELOW", (0, 0), (-1, -1), 0.45, colors.HexColor("#E2E8F0")),
            ]
        )
    )
    story.append(cover_band)
    story.append(Spacer(1, 6))

    header_row = []
    if logo_path and Path(logo_path).exists():
        header_row.append(Image(str(logo_path), width=22 * mm, height=14 * mm))
    else:
        header_row.append(Paragraph(" ", styles["body"]))
    header_row.append(Paragraph(f"<b>{title}</b><br/><font color='{MUTED}'>{subtitle}</font>", styles["title"]))

    metadata = [
        f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        f"Emitido por: {generated_by or 'Sistema'}",
        "Sistema de Checklist de Frota",
    ]
    header_row.append(Paragraph("<br/>".join(metadata), styles["meta"]))

    table = Table([header_row], colWidths=[28 * mm, 116 * mm, 45 * mm])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                ("LINEBELOW", (0, 0), (-1, -1), 0.7, colors.HexColor("#D9E2EF")),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    story.append(table)
    story.append(Spacer(1, 8))
    return story


def _build_cover_page(
    title: str,
    subtitle: str,
    generated_by: str,
    logo_path: str | Path | None,
    styles,
    *,
    landscape_mode: bool,
):
    story = []
    width = 250 * mm if landscape_mode else 180 * mm

    top_band = Table(
        [[Paragraph("Relatório executivo | Sistema de Checklist de Frota", styles["cover_band_large"])]],
        colWidths=[width],
    )
    top_band.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor(MUTED)),
                ("LINEBELOW", (0, 0), (-1, -1), 0.8, colors.HexColor("#D9E2EF")),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    story.append(top_band)
    story.append(Spacer(1, 18))

    if logo_path and Path(logo_path).exists():
        story.append(Image(str(logo_path), width=34 * mm, height=22 * mm))
        story.append(Spacer(1, 14))

    story.append(Paragraph(title, styles["cover_title"]))
    story.append(Spacer(1, 8))
    story.append(Paragraph(subtitle, styles["cover_subtitle"]))
    story.append(Spacer(1, 22))

    story.extend(
        _build_summary_cards(
            [
                ("Emitido por", generated_by or "Sistema"),
                ("Data", datetime.now().strftime("%d/%m/%Y")),
                ("Hora", datetime.now().strftime("%H:%M")),
            ],
            styles,
            compact=False,
        )
    )
    story.append(Spacer(1, 24))

    intro = Table(
        [[Paragraph(
            "Documento gerado para apoio gerencial, rastreabilidade operacional e evidências de manutenção da frota.",
            styles["cover_intro"],
        )]],
        colWidths=[width],
    )
    intro.setStyle(
        TableStyle(
            [
                ("LINEABOVE", (0, 0), (-1, -1), 0.7, colors.HexColor("#D9E2EF")),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    story.append(intro)
    return story


def _build_summary_cards(items: list[tuple[str, str]], styles, *, compact: bool = True):
    if not items:
        return []

    # Mantém os cards dentro da área útil da página em qualquer relatório
    # (principalmente A4 retrato com 4 indicadores).
    max_total_width = 180 * mm
    gap_width = 2.5 * mm
    total_gaps = gap_width * max(0, len(items) - 1)
    card_width = (max_total_width - total_gaps) / len(items)
    min_card_width = 34 * mm if compact else 38 * mm
    card_width = max(min_card_width, card_width)

    cells = []
    for label, value in items:
        accent = _summary_accent_for(label, value)
        card = Table(
            [
                [Paragraph(label, styles["summary_label"])],
                [Paragraph(_safe_paragraph_text(value), styles["summary_value"])],
            ],
            colWidths=[card_width],
        )
        card.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(accent["bg"])),
                    ("LINELEFT", (0, 0), (-1, -1), 1.1, colors.HexColor(accent["border"])),
                    ("LINEBELOW", (0, -1), (-1, -1), 0.35, colors.HexColor("#E2E8F0")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )
        cells.append(card)

    col_widths = [card_width] * len(cells)
    if len(cells) > 1:
        col_widths = [card_width + (gap_width if index < len(cells) - 1 else 0) for index in range(len(cells))]
    wrapper = Table([cells], colWidths=col_widths)
    wrapper.hAlign = "LEFT"
    wrapper.setStyle(
        TableStyle(
            [
                ("BOX", (0, 0), (-1, -1), 0, colors.white),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    return [wrapper]


def _build_signature_block(generated_by: str, styles):
    label = generated_by or "Sistema"
    block = Table(
        [[Paragraph("Emitido por", styles["signature_label"])], [Paragraph(label, styles["signature_value"])]],
        colWidths=[70 * mm],
    )
    block.setStyle(
        TableStyle(
            [
                ("LINEABOVE", (0, 1), (-1, 1), 0.7, colors.HexColor("#94A3B8")),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    return [block]


def _activity_status_stamp(status_text: str, styles):
    normalized = status_text.lower()
    if "instalado" in normalized and "nao" not in normalized and "não" not in normalized:
        accent = {"bg": SEVERITY_GREEN_BG, "border": SEVERITY_GREEN, "text": SEVERITY_GREEN}
    elif "nao instalado" in normalized or "não instalado" in normalized:
        accent = {"bg": SEVERITY_RED_BG, "border": SEVERITY_RED, "text": SEVERITY_RED}
    else:
        accent = {"bg": SEVERITY_YELLOW_BG, "border": SEVERITY_YELLOW, "text": SEVERITY_YELLOW}

    stamp = Table([[Paragraph(_safe_paragraph_text(status_text.upper()), styles["stamp"])]] , colWidths=[42 * mm])
    stamp.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(accent["bg"])),
                ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor(accent["border"])),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    return stamp


def _build_chart_section(columns, rows, styles):
    if _is_nc_macro_dataset(columns):
        return _build_nc_macro_chart_section(rows, styles)

    metric_key, metric_label = _detect_metric_column(columns, rows)
    category_key, _ = _detect_category_column(columns)
    if not metric_key or not category_key:
        return []

    top_rows = sorted(rows, key=lambda row: _to_number(row.get(metric_key)), reverse=True)[:5]
    if not top_rows:
        return []

    max_value = max(_to_number(row.get(metric_key)) for row in top_rows) or 1
    drawing = Drawing(520, 150)
    base_y = 18
    bar_height = 18
    gap = 10
    drawing.add(String(0, 132, f"Gráfico executivo - {metric_label}", fontName="Helvetica-Bold", fontSize=11, fillColor=colors.HexColor(BRAND_BLUE)))

    for index, row in enumerate(top_rows):
        y = 104 - index * (bar_height + gap)
        value = _to_number(row.get(metric_key))
        width = 320 * (value / max_value)
        label = _stringify(row.get(category_key))[:28]
        severity = _severity_for_value(value, max_value)
        drawing.add(String(0, y + 4, label, fontName="Helvetica", fontSize=8.5, fillColor=colors.HexColor(BRAND_BLUE)))
        drawing.add(Rect(132, y, width, bar_height, fillColor=colors.HexColor(severity["color"]), strokeColor=colors.HexColor(severity["color"])))
        drawing.add(String(460, y + 4, str(int(value) if value.is_integer() else value), fontName="Helvetica-Bold", fontSize=8.5, fillColor=colors.HexColor(BRAND_BLUE)))

    wrapper = Table([[drawing]], colWidths=[170 * mm])
    wrapper.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                ("LINEBELOW", (0, 0), (-1, -1), 0.5, colors.HexColor("#D9E2EF")),
                ("LEFTPADDING", (0, 0), (-1, -1), 12),
                ("RIGHTPADDING", (0, 0), (-1, -1), 12),
                ("TOPPADDING", (0, 0), (-1, -1), 12),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
            ]
        )
    )
    return [Paragraph("Gráfico visual", styles["section"]), wrapper]


def _is_nc_macro_dataset(columns: list[tuple[str, str]]) -> bool:
    keys = {key for _, key in columns}
    return {"item_nome", "total_nc", "abertas", "resolvidas"}.issubset(keys)


def _build_nc_macro_chart_section(rows: list[dict], styles):
    if not rows:
        return []

    top_rows = sorted(rows, key=lambda row: _to_number(row.get("total_nc")), reverse=True)[:8]
    if not top_rows:
        return []

    total_series = [max(0.0, _to_number(row.get("total_nc"))) for row in top_rows]
    open_series = [max(0.0, _to_number(row.get("abertas"))) for row in top_rows[:6]]
    resolved_series = [max(0.0, _to_number(row.get("resolvidas"))) for row in top_rows[:6]]
    top_labels = [_truncate_label(_stringify(row.get("item_nome")), 18) for row in top_rows]
    compare_labels = [_truncate_label(_stringify(row.get("item_nome")), 12) for row in top_rows[:6]]

    total_open = sum(max(0.0, _to_number(row.get("abertas"))) for row in rows)
    total_resolved = sum(max(0.0, _to_number(row.get("resolvidas"))) for row in rows)
    base_total = total_open + total_resolved
    if base_total <= 0:
        pie_data = [1.0, 0.0]
        open_percent = 0.0
        resolved_percent = 0.0
    else:
        pie_data = [total_open, total_resolved]
        open_percent = (total_open / base_total) * 100
        resolved_percent = (total_resolved / base_total) * 100

    max_total = max(total_series) if total_series else 1.0
    max_compare = max(open_series + resolved_series) if (open_series or resolved_series) else 1.0
    total_axis_max = max(1.0, max_total * 1.2)
    compare_axis_max = max(1.0, max_compare * 1.2)

    top_bar = Drawing(450, 170)
    top_bar.add(
        String(
            8,
            154,
            "Top não conformidades por volume",
            fontName="Helvetica-Bold",
            fontSize=10.5,
            fillColor=colors.HexColor(BRAND_BLUE),
        )
    )
    total_chart = VerticalBarChart()
    total_chart.x = 34
    total_chart.y = 32
    total_chart.height = 108
    total_chart.width = 396
    total_chart.data = [tuple(total_series)]
    total_chart.categoryAxis.categoryNames = top_labels
    total_chart.categoryAxis.labels.angle = 25
    total_chart.categoryAxis.labels.boxAnchor = "ne"
    total_chart.categoryAxis.labels.fontName = "Helvetica"
    total_chart.categoryAxis.labels.fontSize = 7.4
    total_chart.valueAxis.valueMin = 0
    total_chart.valueAxis.valueMax = total_axis_max
    total_chart.valueAxis.valueStep = max(1, int(round(total_axis_max / 5)))
    total_chart.valueAxis.labels.fontName = "Helvetica"
    total_chart.valueAxis.labels.fontSize = 7.2
    total_chart.bars[0].fillColor = colors.HexColor(PRIMARY_BLUE)
    total_chart.bars[0].strokeColor = colors.HexColor(PRIMARY_BLUE)
    total_chart.bars[0].strokeWidth = 0.2
    top_bar.add(total_chart)

    pie_drawing = Drawing(210, 170)
    pie_drawing.add(
        String(
            8,
            154,
            "Status geral das NC",
            fontName="Helvetica-Bold",
            fontSize=10.0,
            fillColor=colors.HexColor(BRAND_BLUE),
        )
    )
    pie = Pie()
    pie.x = 28
    pie.y = 38
    pie.width = 116
    pie.height = 102
    pie.data = pie_data
    pie.labels = [f"Abertas {open_percent:.0f}%", f"Resolvidas {resolved_percent:.0f}%"]
    pie.sideLabels = True
    pie.simpleLabels = False
    pie.slices.strokeWidth = 0.4
    pie.slices[0].fillColor = colors.HexColor(SEVERITY_RED)
    pie.slices[1].fillColor = colors.HexColor(SEVERITY_GREEN)
    pie_drawing.add(pie)
    pie_drawing.add(
        String(
            8,
            18,
            f"Abertas: {int(total_open)} | Resolvidas: {int(total_resolved)}",
            fontName="Helvetica-Bold",
            fontSize=8.2,
            fillColor=colors.HexColor(BRAND_BLUE),
        )
    )

    compare_drawing = Drawing(210, 170)
    compare_drawing.add(
        String(
            8,
            154,
            "Abertas x resolvidas por NC",
            fontName="Helvetica-Bold",
            fontSize=10.0,
            fillColor=colors.HexColor(BRAND_BLUE),
        )
    )
    compare_chart = VerticalBarChart()
    compare_chart.x = 18
    compare_chart.y = 40
    compare_chart.height = 102
    compare_chart.width = 176
    compare_chart.data = [tuple(open_series), tuple(resolved_series)]
    compare_chart.categoryAxis.categoryNames = compare_labels
    compare_chart.categoryAxis.labels.angle = 20
    compare_chart.categoryAxis.labels.boxAnchor = "ne"
    compare_chart.categoryAxis.labels.fontName = "Helvetica"
    compare_chart.categoryAxis.labels.fontSize = 6.9
    compare_chart.valueAxis.valueMin = 0
    compare_chart.valueAxis.valueMax = compare_axis_max
    compare_chart.valueAxis.valueStep = max(1, int(round(compare_axis_max / 5)))
    compare_chart.valueAxis.labels.fontName = "Helvetica"
    compare_chart.valueAxis.labels.fontSize = 6.8
    compare_chart.bars[0].fillColor = colors.HexColor(SEVERITY_RED)
    compare_chart.bars[0].strokeColor = colors.HexColor(SEVERITY_RED)
    compare_chart.bars[1].fillColor = colors.HexColor(SEVERITY_GREEN)
    compare_chart.bars[1].strokeColor = colors.HexColor(SEVERITY_GREEN)
    compare_chart.bars[0].strokeWidth = 0.2
    compare_chart.bars[1].strokeWidth = 0.2
    compare_drawing.add(compare_chart)
    legend = Legend()
    legend.x = 54
    legend.y = 14
    legend.fontName = "Helvetica"
    legend.fontSize = 7.6
    legend.colorNamePairs = [
        (colors.HexColor(SEVERITY_RED), "Abertas"),
        (colors.HexColor(SEVERITY_GREEN), "Resolvidas"),
    ]
    compare_drawing.add(legend)

    top_wrapper = _chart_card(top_bar, 170 * mm)
    bottom_wrapper = Table(
        [[_chart_card(pie_drawing, 84 * mm), _chart_card(compare_drawing, 84 * mm)]],
        colWidths=[84 * mm, 84 * mm],
    )
    bottom_wrapper.setStyle(
        TableStyle(
            [
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    return [Paragraph("Painel gráfico executivo", styles["section"]), top_wrapper, Spacer(1, 6), bottom_wrapper]


def _chart_card(drawing: Drawing, width: float) -> Table:
    card = Table([[drawing]], colWidths=[width])
    card.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                ("BOX", (0, 0), (-1, -1), 0.55, colors.HexColor("#D9E2EF")),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    return card


def _truncate_label(value: str, max_len: int) -> str:
    text = (value or "").strip()
    if len(text) <= max_len:
        return text
    return text[: max(0, max_len - 3)].rstrip() + "..."


def _build_top_five_section(columns, rows, styles):
    metric_key, metric_label = _detect_metric_column(columns, rows)
    category_key, category_label = _detect_category_column(columns)
    if not metric_key or not category_key:
        return []

    top_rows = sorted(rows, key=lambda row: _to_number(row.get(metric_key)), reverse=True)[:5]
    if not top_rows:
        return []

    table_data = [[
        Paragraph("Posição", styles["table_header"]),
        Paragraph(category_label, styles["table_header"]),
        Paragraph(metric_label, styles["table_header"]),
        Paragraph("Prioridade", styles["table_header"]),
    ]]
    row_backgrounds = []
    for index, row in enumerate(top_rows, start=1):
        severity = _severity_for_value(_to_number(row.get(metric_key)), _to_number(top_rows[0].get(metric_key)) or 1)
        row_backgrounds.append(colors.HexColor(severity["bg"]))
        table_data.append(
            [
                Paragraph(str(index), styles["table_cell"]),
                Paragraph(_safe_paragraph_text(_stringify(row.get(category_key))), styles["table_cell"]),
                Paragraph(_safe_paragraph_text(_stringify(row.get(metric_key))), styles["table_cell"]),
                Paragraph(severity["label"], styles["table_cell"]),
            ]
        )

    table = Table(table_data, colWidths=[20 * mm, 88 * mm, 28 * mm, 30 * mm], repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(PRIMARY_BLUE)),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("LINEBELOW", (0, 0), (-1, -1), 0.35, colors.HexColor("#E2E8F0")),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ]
        )
    )
    for index, background in enumerate(row_backgrounds, start=1):
        table.setStyle(TableStyle([("BACKGROUND", (0, index), (-1, index), background)]))
    return [Paragraph("Ranking Top 5", styles["section"]), table]


def _build_executive_conclusion(columns, rows, styles, period_label: str | None):
    metric_key, metric_label = _detect_metric_column(columns, rows)
    category_key, category_label = _detect_category_column(columns)
    period_text = period_label or f"base consolidada até {datetime.now().strftime('%d/%m/%Y')}"
    if not rows or not metric_key or not category_key:
        conclusion = (
            f"Conclusão executiva: não foram identificados dados suficientes para análise consolidada no período {period_text}."
        )
        recommendation = "Prioridade sugerida: manter monitoramento e consolidar nova base antes de ações corretivas amplas."
        severity = {"bg": LIGHT_BG, "border": "#D9E2EF"}
    else:
        sorted_rows = sorted(rows, key=lambda row: _to_number(row.get(metric_key)), reverse=True)
        top_row = sorted_rows[0]
        total_metric = sum(_to_number(row.get(metric_key)) for row in rows)
        average_metric = total_metric / max(len(rows), 1)
        top_value = _to_number(top_row.get(metric_key))
        severity = _severity_for_value(top_value, top_value)
        priority_text = _overall_priority_label(columns, rows)
        conclusion = (
            f"Conclusão executiva: no período {period_text}, o indicador principal '{metric_label}' somou "
            f"{int(total_metric) if total_metric.is_integer() else round(total_metric, 2)}. "
            f"O maior destaque ficou em {category_label.lower()} '{_stringify(top_row.get(category_key))}', com "
            f"{int(_to_number(top_row.get(metric_key))) if _to_number(top_row.get(metric_key)).is_integer() else round(_to_number(top_row.get(metric_key)), 2)}. "
            f"A média por registro foi de {round(average_metric, 2)}."
        )
        recommendation = (
            f"Prioridade sugerida: {priority_text.lower()}. Direcionar a atuação imediata sobre "
            f"{category_label.lower()} '{_stringify(top_row.get(category_key))}' e acompanhar a evolução das demais ocorrências no mesmo período."
        )

    box = Table(
        [
            [Paragraph(_safe_paragraph_text(conclusion), styles["body"])],
            [Paragraph(_safe_paragraph_text(recommendation), styles["body"])],
        ],
        colWidths=[170 * mm],
    )
    box.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(severity["bg"])),
                ("LINELEFT", (0, 0), (-1, -1), 1.1, colors.HexColor(severity["border"])),
                ("LINEBELOW", (0, -1), (-1, -1), 0.35, colors.HexColor("#E2E8F0")),
                ("LEFTPADDING", (0, 0), (-1, -1), 12),
                ("RIGHTPADDING", (0, 0), (-1, -1), 12),
                ("TOPPADDING", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ]
        )
    )
    return [Paragraph("Conclusão executiva", styles["section"]), box]


def _overall_priority_label(columns, rows) -> str:
    metric_key, _ = _detect_metric_column(columns, rows)
    if not rows or not metric_key:
        return "Monitoramento preventivo"
    max_value = max(_to_number(row.get(metric_key)) for row in rows) or 1
    severity = _severity_for_value(max_value, max_value)
    if severity["label"] == "Alta":
        return "Alta prioridade"
    if severity["label"] == "Moderada":
        return "Prioridade moderada"
    return "Cenário controlado"


def _summary_accent_for(label: str, value: str):
    text = f"{label} {value}".lower()
    if "semaforo executivo" in text or "semáforo executivo" in text:
        if "alta" in text:
            return {"bg": SEVERITY_RED_BG, "border": SEVERITY_RED}
        if "moderada" in text:
            return {"bg": SEVERITY_YELLOW_BG, "border": SEVERITY_YELLOW}
        return {"bg": SEVERITY_GREEN_BG, "border": SEVERITY_GREEN}
    return {"bg": "#FFFFFF", "border": "#D9E2EF"}


def _severity_for_value(value: float, max_value: float):
    ratio = 0 if max_value <= 0 else value / max_value
    if value <= 0:
        return {"label": "Controlada", "color": SEVERITY_GREEN, "bg": SEVERITY_GREEN_BG, "border": SEVERITY_GREEN}
    if ratio >= 0.67 or value >= 10:
        return {"label": "Alta", "color": SEVERITY_RED, "bg": SEVERITY_RED_BG, "border": SEVERITY_RED}
    if ratio >= 0.34 or value >= 4:
        return {"label": "Moderada", "color": SEVERITY_YELLOW, "bg": SEVERITY_YELLOW_BG, "border": SEVERITY_YELLOW}
    return {"label": "Controlada", "color": SEVERITY_GREEN, "bg": SEVERITY_GREEN_BG, "border": SEVERITY_GREEN}


def _detect_metric_column(columns, rows):
    preferred = ["total_nc", "nc", "abertas", "resolvidas"]
    labels = dict(columns)
    for key in preferred:
        if any(column_key == key for _, column_key in columns):
            return key, next(label for label, column_key in columns if column_key == key)
    for label, key in columns:
        if rows and any(_is_number(row.get(key)) for row in rows):
            return key, label
    return None, None


def _detect_category_column(columns):
    preferred = ["item_nome", "item", "frota", "modelo"]
    for key in preferred:
        for label, column_key in columns:
            if column_key == key:
                return column_key, label
    if columns:
        return columns[0][1], columns[0][0]
    return None, None


def _is_number(value) -> bool:
    try:
        float(value)
        return True
    except (TypeError, ValueError):
        return False


def _to_number(value) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _page_frame_callback(title: str, subtitle: str, generated_by: str, logo_path: str | Path | None):
    def draw(canvas, document):
        _draw_page_frame(canvas, document, generated_by, title, subtitle, logo_path)

    return draw


def _draw_page_frame(
    canvas,
    document,
    generated_by: str,
    title: str | None = None,
    subtitle: str | None = None,
    logo_path: str | Path | None = None,
):
    if canvas.getPageNumber() > 1:
        _draw_report_page_header(canvas, document, title or "Relatório corporativo", subtitle or "", generated_by, logo_path)
    _draw_footer(canvas, document, generated_by)


def _draw_report_page_header(
    canvas,
    document,
    title: str,
    subtitle: str,
    generated_by: str,
    logo_path: str | Path | None,
):
    canvas.saveState()
    page_width, page_height = document.pagesize
    content_width = page_width - document.leftMargin - document.rightMargin
    left = document.leftMargin
    right = left + content_width
    top = page_height - 12 * mm
    line_y = top - 7 * mm

    canvas.setFillColor(colors.HexColor(MUTED))
    canvas.setFont("Helvetica-Bold", 8)
    canvas.drawString(left, top, "Sistema de Checklist de Frota | Relatório corporativo")
    canvas.setStrokeColor(colors.HexColor("#D9E2EF"))
    canvas.setLineWidth(0.45)
    canvas.line(left, line_y, right, line_y)

    logo_bottom = line_y - 20 * mm
    logo_width = 24 * mm
    logo_height = 16 * mm
    logo_file = Path(logo_path) if logo_path else None
    if logo_file and logo_file.exists():
        try:
            canvas.drawImage(
                str(logo_file),
                left,
                logo_bottom + 2 * mm,
                width=logo_width,
                height=logo_height,
                preserveAspectRatio=True,
                mask="auto",
            )
        except Exception:
            pass

    text_left = left + 30 * mm
    title_y = line_y - 7 * mm
    canvas.setFillColor(colors.HexColor("#0B1220"))
    canvas.setFont("Helvetica-Bold", 11)
    canvas.drawString(text_left, title_y, _canvas_text(title, 68))

    canvas.setFillColor(colors.HexColor(MUTED))
    canvas.setFont("Helvetica", 10)
    canvas.drawString(text_left, title_y - 6 * mm, _canvas_text(subtitle, 82))

    metadata = [
        f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        f"Emitido por: {generated_by or 'Sistema'}",
        "Sistema de Checklist de Frota",
    ]
    canvas.setFont("Helvetica", 7)
    meta_y = title_y - 1 * mm
    for index, line in enumerate(metadata):
        canvas.drawRightString(right, meta_y - index * 4.2 * mm, _canvas_text(line, 42))

    canvas.setStrokeColor(colors.HexColor("#D9E2EF"))
    canvas.setLineWidth(0.7)
    canvas.line(left, logo_bottom, right, logo_bottom)
    canvas.restoreState()


def _canvas_text(value: str | None, max_chars: int) -> str:
    text = str(value or "")
    if len(text) <= max_chars:
        return text
    return text[: max(0, max_chars - 3)].rstrip() + "..."


def _draw_footer(canvas, document, generated_by: str):
    canvas.saveState()
    footer_y = 9 * mm
    width = document.pagesize[0] - document.leftMargin - document.rightMargin
    canvas.setStrokeColor(colors.HexColor("#CBD5E1"))
    canvas.line(document.leftMargin, footer_y + 4 * mm, document.leftMargin + width, footer_y + 4 * mm)
    canvas.setFillColor(colors.HexColor(MUTED))
    canvas.setFont("Helvetica", 8)
    canvas.drawString(document.leftMargin, footer_y, "Sistema de Checklist de Frota")
    canvas.drawCentredString(document.leftMargin + width / 2, footer_y, f"Emitido por: {generated_by or 'Sistema'}")
    canvas.drawRightString(document.leftMargin + width, footer_y, f"Página {canvas.getPageNumber()}")
    canvas.restoreState()


def _key_value_table(rows: list[list[str]], styles):
    table = Table(
        [[Paragraph(f"<b>{label}</b>", styles["body"]), Paragraph(_safe_paragraph_text(value), styles["body"])] for label, value in rows],
        colWidths=[48 * mm, 122 * mm],
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor("#FBFDFF")]),
                ("LINEBELOW", (0, 0), (-1, -1), 0.35, colors.HexColor("#E2E8F0")),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    return table


def _standard_table_style() -> TableStyle:
    return TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(PRIMARY_BLUE)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("LINEBELOW", (0, 0), (-1, -1), 0.35, colors.HexColor("#E2E8F0")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor(LIGHT_BG)]),
            ("LEFTPADDING", (0, 0), (-1, -1), 7),
            ("RIGHTPADDING", (0, 0), (-1, -1), 7),
            ("TOPPADDING", (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]
    )


def _append_occurrence_evidence_section(
    story: list,
    occurrences: list[dict],
    occurrence_images: dict[int, dict[str, bytes | None]],
    styles,
    *,
    section_title: str,
    page_break_before: bool = False,
    include_resolution_fields: bool = True,
    include_part_fields: bool = True,
) -> None:
    if page_break_before:
        story.append(PageBreak())
    story.append(Paragraph(section_title, styles["section"]))
    story.append(Spacer(1, 6))

    for index, item in enumerate(occurrences, start=1):
        if index > 1:
            story.append(PageBreak())

        vehicle = item.get("veiculo") or {}
        user = item.get("usuario") or {}
        resolved_by = item.get("resolved_by") or {}
        status_text = "Resolvida" if item.get("resolvido") else "Aberta"

        story.append(
            Paragraph(
                _safe_paragraph_text(
                    f"Ocorrência {index} - {vehicle.get('frota') or '-'} - {item.get('item_nome') or '-'}"
                ),
                styles["section"],
            )
        )
        story.append(Spacer(1, 4))
        details = [
            ["Equipamento", vehicle.get("frota") or "-"],
            ["Placa", vehicle.get("placa") or "-"],
            ["Modelo", vehicle.get("modelo") or "-"],
            ["Item", item.get("item_nome") or "-"],
            ["Status", status_text],
            ["Motorista", user.get("nome") or "-"],
            ["Abertura", _format_datetime(item.get("created_at"))],
        ]
        if include_resolution_fields and (item.get("resolvido") or item.get("data_resolucao") or resolved_by.get("nome")):
            details.extend(
                [
                    ["Resolução", _format_datetime(item.get("data_resolucao"))],
                    ["Resolvido por", resolved_by.get("nome") or "-"],
                ]
            )
        if include_part_fields:
            codigo_peca = str(item.get("codigo_peca") or "").strip()
            if codigo_peca:
                details.append(["Código da peça", codigo_peca])
            descricao_peca = str(item.get("descricao_peca") or "").strip()
            if descricao_peca:
                details.append(["Descrição da peça", descricao_peca])
        observacao = str(item.get("observacao") or "").strip()
        if observacao:
            details.append(["Observação", observacao])
        story.append(
            _key_value_table(
                details,
                styles,
            )
        )
        story.append(Spacer(1, 8))

        images = occurrence_images.get(item.get("id"), {})
        before = images.get("before")
        after = images.get("after")
        image_block = [
            _reportlab_image(before, 86 * mm, 76 * mm) if before else Paragraph("Sem foto de origem", styles["muted_box"]),
            _reportlab_image(after, 86 * mm, 76 * mm) if after else Paragraph("Sem foto de resolução", styles["muted_box"]),
        ]
        image_table = Table(
            [
                [
                    Paragraph("Foto de origem", styles["section"]),
                    Paragraph("Foto de resolução", styles["section"]),
                ],
                image_block,
            ],
            colWidths=[89 * mm, 89 * mm],
        )
        image_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.white),
                    ("LINEBELOW", (0, 0), (-1, 0), 0.35, colors.HexColor("#CBD5E1")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E2E8F0")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        story.append(image_table)


def _reportlab_image(raw_bytes: bytes, width: float, height: float):
    buffer = io.BytesIO(raw_bytes)
    image = Image(buffer, width=width, height=height, kind="proportional")
    image.hAlign = "CENTER"
    return image


def _styles():
    base = getSampleStyleSheet()
    return {
        "title": ParagraphStyle(
            "TitleBlock",
            parent=base["BodyText"],
            fontName="Helvetica",
            fontSize=15,
            leading=19,
            textColor=colors.HexColor(BRAND_BLUE),
        ),
        "meta": ParagraphStyle(
            "MetaBlock",
            parent=base["BodyText"],
            fontName="Helvetica",
            fontSize=8.5,
            leading=11,
            textColor=colors.HexColor(MUTED),
            alignment=2,
        ),
        "body": ParagraphStyle(
            "Body",
            parent=base["BodyText"],
            fontName="Helvetica",
            fontSize=9.5,
            leading=12,
            textColor=colors.HexColor(BRAND_BLUE),
        ),
        "section": ParagraphStyle(
            "Section",
            parent=base["Heading3"],
            fontName="Helvetica-Bold",
            fontSize=11,
            textColor=colors.HexColor(BRAND_BLUE),
            spaceAfter=4,
        ),
        "table_header": ParagraphStyle(
            "TableHeader",
            parent=base["BodyText"],
            fontName="Helvetica-Bold",
            fontSize=8.5,
            leading=11,
            textColor=colors.white,
        ),
        "table_cell": ParagraphStyle(
            "TableCell",
            parent=base["BodyText"],
            fontName="Helvetica",
            fontSize=8.3,
            leading=10,
            textColor=colors.HexColor(BRAND_BLUE),
        ),
        "muted_box": ParagraphStyle(
            "MutedBox",
            parent=base["BodyText"],
            fontName="Helvetica",
            fontSize=9.5,
            alignment=1,
            textColor=colors.HexColor(MUTED),
        ),
        "cover_band": ParagraphStyle(
            "CoverBand",
            parent=base["BodyText"],
            fontName="Helvetica-Bold",
            fontSize=10,
            leading=12,
            textColor=colors.HexColor(MUTED),
        ),
        "cover_band_large": ParagraphStyle(
            "CoverBandLarge",
            parent=base["BodyText"],
            fontName="Helvetica-Bold",
            fontSize=12,
            leading=14,
            textColor=colors.HexColor(MUTED),
        ),
        "cover_title": ParagraphStyle(
            "CoverTitle",
            parent=base["Heading1"],
            fontName="Helvetica-Bold",
            fontSize=22,
            leading=26,
            textColor=colors.HexColor(BRAND_BLUE),
        ),
        "cover_subtitle": ParagraphStyle(
            "CoverSubtitle",
            parent=base["BodyText"],
            fontName="Helvetica",
            fontSize=12,
            leading=16,
            textColor=colors.HexColor(MUTED),
        ),
        "cover_intro": ParagraphStyle(
            "CoverIntro",
            parent=base["BodyText"],
            fontName="Helvetica",
            fontSize=11,
            leading=15,
            textColor=colors.HexColor(BRAND_BLUE),
        ),
        "summary_label": ParagraphStyle(
            "SummaryLabel",
            parent=base["BodyText"],
            fontName="Helvetica-Bold",
            fontSize=8.5,
            leading=10,
            textColor=colors.HexColor(MUTED),
        ),
        "summary_value": ParagraphStyle(
            "SummaryValue",
            parent=base["BodyText"],
            fontName="Helvetica-Bold",
            fontSize=15,
            leading=18,
            textColor=colors.HexColor(BRAND_BLUE),
        ),
        "signature_label": ParagraphStyle(
            "SignatureLabel",
            parent=base["BodyText"],
            fontName="Helvetica",
            fontSize=8.5,
            leading=10,
            textColor=colors.HexColor(MUTED),
        ),
        "signature_value": ParagraphStyle(
            "SignatureValue",
            parent=base["BodyText"],
            fontName="Helvetica-Bold",
            fontSize=10,
            leading=12,
            textColor=colors.HexColor(BRAND_BLUE),
        ),
        "stamp": ParagraphStyle(
            "Stamp",
            parent=base["BodyText"],
            fontName="Helvetica-Bold",
            fontSize=9,
            leading=11,
            alignment=1,
            textColor=colors.HexColor(BRAND_BLUE),
        ),
    }


def _stringify(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Sim" if value else "Não"
    return str(value)


def _activity_item_status_text(value: str | None) -> str:
    return {
        "PENDENTE": "Pendente",
        "INSTALADO": "Instalado",
        "NAO_INSTALADO": "Não instalado",
    }.get(value or "", value or "-")


def _format_datetime(value: str | None) -> str:
    if not value:
        return "-"
    try:
        text = value.replace("Z", "")
        return datetime.fromisoformat(text).strftime("%d/%m/%Y %H:%M")
    except ValueError:
        return value


def _safe_paragraph_text(value: str) -> str:
    return (
        value.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace("\n", "<br/>")
    )

