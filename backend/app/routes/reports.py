import csv
from io import BytesIO, StringIO
from datetime import datetime, time

from flask import Blueprint, g, request, send_file
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import func, or_
from app.extensions import db
from app.models import ChecklistItem, Vehicle, Checklist
from app.services.auth_service import auth_required, user_has_management_access
from app.services.bi_contract_service import build_readonly_bi_contract
from app.services.maintenance_intelligence_service import build_maintenance_intelligence_overview
from app.services.report_service import (
    MASTER_BASE_EXPORT_COLUMNS,
    build_dashboard_summary,
    build_management_master_export,
    build_management_master_base,
    build_productivity_report,
)
from app.services.audit_service import record_event
from app.utils.responses import api_response
from app.utils.filters import apply_item_search
from app.routes.non_conformities import NCStatus

bp = Blueprint("reports", __name__, url_prefix="/relatorios")

def _parse_date(value: str | None, end_of_day: bool = False):
    if not value:
        return None
    try:
        dt = datetime.fromisoformat(str(value).strip())
        if len(str(value).strip()) <= 10:
            return datetime.combine(dt.date(), time.max if end_of_day else time.min)
        return dt
    except ValueError:
        return None


@bp.get("/dashboard")
@auth_required
def get_dashboard_report():
    """Resumo executivo do dashboard (mantido para desktop e web)."""
    return api_response(True, data=build_dashboard_summary())


@bp.get("/produtividade")
@auth_required
def get_productivity_dashboard():
    """Relatório consolidado de produtividade por usuário."""
    return api_response(True, data=build_productivity_report())


@bp.get("/manutencao-executivo")
@auth_required
def get_maintenance_executive_report():
    if not user_has_management_access(g.current_user):
        return api_response(False, error="Somente admin ou gestor podem consultar este relatorio.", status_code=403)
    return api_response(True, data=build_maintenance_intelligence_overview())


@bp.get("/bi/contrato")
@auth_required
def get_bi_readonly_contract():
    if not user_has_management_access(g.current_user):
        return api_response(False, error="Somente admin ou gestor podem consultar o contrato de BI.", status_code=403)
    return api_response(True, data=build_readonly_bi_contract())


def _management_master_filters() -> dict:
    active_only = request.args.get("ativos", "true").strip().lower() not in {"false", "0", "nao"}
    return {
        "page": request.args.get("pagina", default=1, type=int),
        "page_size": request.args.get("tamanho_pagina", default=50, type=int),
        "family_code": request.args.get("familia"),
        "vehicle_id": request.args.get("equipamento", type=int),
        "location_id": request.args.get("local", type=int),
        "status": request.args.get("status"),
        "source_type": request.args.get("origem"),
        "search": request.args.get("busca"),
        "date_from": request.args.get("data_inicial"),
        "date_to": request.args.get("data_final"),
        "active_only": active_only,
    }


def _management_export_filters() -> dict:
    filters = _management_master_filters()
    filters.pop("page", None)
    filters.pop("page_size", None)
    return filters


@bp.get("/base-mestre")
@auth_required
def get_management_master_base():
    if not user_has_management_access(g.current_user):
        return api_response(False, error="Somente admin ou gestor podem consultar a Base Mestre.", status_code=403)

    try:
        data = build_management_master_base(**_management_master_filters())
    except ValueError as exc:
        return api_response(False, error=str(exc), status_code=400)
    return api_response(True, data=data)


@bp.get("/base-mestre/exportar")
@auth_required
def export_management_master_base():
    if not user_has_management_access(g.current_user):
        return api_response(False, error="Somente admin ou gestor podem exportar a Base Mestre.", status_code=403)

    export_format = (request.args.get("formato") or "json").strip().lower()
    if export_format == "excel":
        export_format = "xlsx"
    if export_format not in {"json", "csv", "xlsx"}:
        return api_response(False, error="Formato invalido. Use json, csv ou xlsx.", status_code=400)

    try:
        payload = build_management_master_export(**_management_export_filters())
    except ValueError as exc:
        return api_response(False, error=str(exc), status_code=400)

    record_event(
        user_id=g.current_user.id,
        entity_type="MANAGEMENT_MASTER_BASE",
        entity_id=0,
        action="EXPORT",
        new_value=f"Formato: {export_format}; registros: {payload['exported']}; total: {payload['total']}",
    )
    db.session.commit()

    if export_format == "json":
        return api_response(True, data=payload)

    if export_format == "csv":
        stream = StringIO(newline="")
        writer = csv.DictWriter(stream, fieldnames=MASTER_BASE_EXPORT_COLUMNS)
        writer.writeheader()
        writer.writerows(payload["items"])
        buffer = BytesIO(stream.getvalue().encode("utf-8-sig"))
        return send_file(
            buffer,
            mimetype="text/csv; charset=utf-8",
            as_attachment=True,
            download_name="base_mestre_pcm.csv",
        )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Base Mestre"
    sheet.append(list(MASTER_BASE_EXPORT_COLUMNS))
    for row in payload["items"]:
        sheet.append([row.get(column) for column in MASTER_BASE_EXPORT_COLUMNS])
    for index, column in enumerate(MASTER_BASE_EXPORT_COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = min(max(len(column) + 2, 12), 32)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="base_mestre_pcm.xlsx",
    )

@bp.get("/macro")
@auth_required
def get_macro_report():
    """Relatório macro: consolidado agrupado por nome do item (Ponto 7: Métricas)."""
    item_principal = func.coalesce(ChecklistItem.item_principal, ChecklistItem.item_nome)
    query = db.session.query(
        item_principal.label("item_nome"),
        func.count(ChecklistItem.id).label("total_nc"),
        func.sum(func.cast(ChecklistItem.resolvido, db.Integer)).label("resolvidas")
    ).filter(ChecklistItem.status == NCStatus.TYPE_NC)

    modulo = request.args.get("modulo")
    if modulo in ("cavalo", "carreta"):
        query = query.join(Checklist).join(Vehicle).filter(Vehicle.tipo == modulo)

    results = query.group_by(item_principal).order_by(func.count(ChecklistItem.id).desc()).all()

    data = [{
        "item_nome": r.item_nome,
        "total_nc": r.total_nc,
        "resolvidas": int(r.resolvidas or 0),
        "abertas": r.total_nc - int(r.resolvidas or 0)
    } for r in results]

    return api_response(True, data=data)

@bp.get("/micro")
@auth_required
def get_micro_report():
    """Relatório micro: ranking de equipamentos com mais ocorrências."""
    query = db.session.query(
        Vehicle.id.label("vehicle_id"),
        Vehicle.frota,
        Vehicle.placa,
        Vehicle.modelo,
        Vehicle.tipo,
        func.count(ChecklistItem.id).label("total_nc"),
        func.max(Checklist.created_at).label("ultimo_checklist")
    ).join(Checklist, Checklist.vehicle_id == Vehicle.id)\
     .outerjoin(ChecklistItem, (ChecklistItem.checklist_id == Checklist.id) & (ChecklistItem.status == NCStatus.TYPE_NC))

    if request.args.get("ativos") == "true":
        query = query.filter(Vehicle.ativo == True)

    results = query.group_by(Vehicle.id).order_by(func.count(ChecklistItem.id).desc()).all()

    data = [{
        "vehicle_id": r.vehicle_id,
        "frota": r.frota,
        "placa": r.placa,
        "modelo": r.modelo,
        "tipo": r.tipo,
        "total_nc": r.total_nc,
        "ultimo_checklist": r.ultimo_checklist.isoformat() if r.ultimo_checklist else None
    } for r in results]

    return api_response(True, data=data)

@bp.get("/item")
@auth_required
def get_item_report():
    """Consulta detalhada de NCs usando o filtro centralizado."""
    query = ChecklistItem.query.join(Checklist).join(Vehicle).filter(ChecklistItem.status == NCStatus.TYPE_NC)

    # Uso da lógica centralizada de busca
    query = apply_item_search(query, ChecklistItem, request.args.get("item"))

    nc_status = request.args.get("status_nc") or request.args.get("nc_status")
    if nc_status == "abertas":
        query = query.filter(ChecklistItem.resolvido == False)
    elif nc_status == "resolvidas":
        query = query.filter(ChecklistItem.resolvido == True)

    modulo = request.args.get("modulo")
    if modulo in ("cavalo", "carreta"):
        query = query.filter(Vehicle.tipo == modulo)

    date_from = _parse_date(request.args.get("data_de") or request.args.get("date_from"))
    date_to = _parse_date(request.args.get("data_ate") or request.args.get("date_to"), end_of_day=True)
    data_base = request.args.get("data_base", "criacao")
    date_col = ChecklistItem.data_resolucao if data_base == "resolucao" else ChecklistItem.created_at

    if data_base == "resolucao":
        query = query.filter(ChecklistItem.data_resolucao.isnot(None))

    if date_from:
        query = query.filter(date_col >= date_from)
    if date_to:
        query = query.filter(date_col <= date_to)

    results = query.order_by(ChecklistItem.created_at.desc()).all()
    return api_response(True, data=[item.to_dict() for item in results])
