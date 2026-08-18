from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def _number(value):
    return float(value or 0)


def export_purchase_report_xlsx(data: dict, output_path: str | Path, generated_by: str = "") -> Path:
    path = Path(output_path)
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Resumo"
    header_fill = PatternFill("solid", fgColor="0B3D91")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(size=14, bold=True, color="0B3D91")
    summary_sheet["A1"] = "RELATÓRIO DE COMPRAS"
    summary_sheet["A1"].font = title_font
    summary_sheet["A2"] = "Gerado por"
    summary_sheet["B2"] = generated_by or "SIS MMP"
    summary_sheet["A3"] = "Período"
    summary_sheet["B3"] = data.get("period_label") or "Todos os registros"
    summary_sheet.append([])
    summary_sheet.append(["INDICADOR", "VALOR"])
    for cell in summary_sheet[5]:
        cell.fill = header_fill
        cell.font = header_font
    labels = {
        "processes": "Processos",
        "items": "Itens",
        "requested_quantity": "Quantidade solicitada",
        "ordered_quantity": "Quantidade em PC",
        "invoiced_quantity": "Quantidade em NF",
        "received_quantity": "Quantidade recebida",
        "remaining_quantity": "Saldo pendente",
    }
    for key, label in labels.items():
        summary_sheet.append([label, _number(data.get("summary", {}).get(key))])
    for title, key in (("POR STATUS", "by_status"), ("POR TIPO", "by_type"), ("POR PROVEDOR", "by_provider")):
        summary_sheet.append([])
        summary_sheet.append([title, "QUANTIDADE"])
        for cell in summary_sheet[summary_sheet.max_row]:
            cell.fill = header_fill
            cell.font = header_font
        for name, value in (data.get(key) or {}).items():
            amount = value.get("items", 0) if isinstance(value, dict) else value
            summary_sheet.append([name, amount])
    for column, width in {"A": 30, "B": 22}.items():
        summary_sheet.column_dimensions[column].width = width

    items_sheet = workbook.create_sheet("Itens")
    columns = ["SC", "TIPO", "DESCRIÇÃO", "MÓDULO", "EQUIPAMENTO", "STATUS", "PRÓXIMA AÇÃO", "SOLICITADO", "PC", "NF", "RECEBIDO", "SALDO"]
    items_sheet.append(columns)
    for cell in items_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for row in data.get("items", []):
        pcs = ", ".join(str(item.get("pc_number") or "") for item in row.get("purchase_orders", []) if item.get("pc_number"))
        invoices = ", ".join(str(invoice.get("invoice_number") or "") for item in row.get("purchase_orders", []) for invoice in item.get("invoices", []) if invoice.get("invoice_number"))
        items_sheet.append([
            row.get("sc_number"), row.get("item_type"), row.get("description_raw"), row.get("module"), row.get("equipment_raw"),
            row.get("item_status"), row.get("next_action"), _number(row.get("requested_quantity")), pcs, invoices,
            _number(row.get("received_quantity")), _number(row.get("remaining_quantity")),
        ])
    for column, width in {"A": 16, "B": 13, "C": 34, "D": 15, "E": 18, "F": 24, "G": 20, "H": 13, "I": 20, "J": 18, "K": 13, "L": 13}.items():
        items_sheet.column_dimensions[column].width = width
    items_sheet.freeze_panes = "A2"
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def export_purchase_report_pdf(data: dict, output_path: str | Path, generated_by: str = "") -> Path:
    path = Path(output_path)
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="ReportSmall", parent=styles["Normal"], fontName="Helvetica", fontSize=7, leading=9, alignment=TA_LEFT))
    styles.add(ParagraphStyle(name="ReportTitle", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=16, textColor=colors.HexColor("#0B3D91"), spaceAfter=6))
    styles.add(ParagraphStyle(name="ReportHeading", parent=styles["Heading2"], fontName="Helvetica-Bold", fontSize=9, textColor=colors.HexColor("#0B3D91"), spaceBefore=8, spaceAfter=4))
    document = SimpleDocTemplate(str(path), pagesize=landscape(A4), leftMargin=12 * mm, rightMargin=12 * mm, topMargin=12 * mm, bottomMargin=12 * mm)
    story = [Paragraph("RELATÓRIO DE COMPRAS", styles["ReportTitle"]), Paragraph(f"Gerado por: {generated_by or 'SIS MMP'} · Período: {data.get('period_label') or 'Todos os registros'}", styles["ReportSmall"]), Spacer(1, 6)]
    summary = data.get("summary", {})
    metric_rows = [["PROCESSOS", "ITENS", "SOLICITADO", "EM PC", "EM NF", "RECEBIDO", "SALDO"], [str(summary.get("processes", 0)), str(summary.get("items", 0)), str(_number(summary.get("requested_quantity"))), str(_number(summary.get("ordered_quantity"))), str(_number(summary.get("invoiced_quantity"))), str(_number(summary.get("received_quantity"))), str(_number(summary.get("remaining_quantity")))]]
    metrics = Table(metric_rows, colWidths=[35 * mm] * 7)
    metrics.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0B3D91")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("ALIGN", (0, 0), (-1, -1), "CENTER"), ("GRID", (0, 0), (-1, -1), .25, colors.HexColor("#CBD5E1")), ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#F1F5F9")), ("FONTSIZE", (0, 0), (-1, -1), 7)]))
    story.extend([metrics, Spacer(1, 8), Paragraph("ITENS DO PROCESSO", styles["ReportHeading"])])
    table_data = [["SC", "TIPO", "DESCRIÇÃO", "MÓDULO", "STATUS", "AÇÃO", "SOLIC.", "RECEB.", "SALDO"]]
    for row in data.get("items", [])[:1000]:
        table_data.append([row.get("sc_number") or "-", row.get("item_type") or "-", str(row.get("description_raw") or "-")[:46], row.get("module") or "-", row.get("item_status") or "-", row.get("next_action") or "-", str(_number(row.get("requested_quantity"))), str(_number(row.get("received_quantity"))), str(_number(row.get("remaining_quantity")))])
    items_table = Table(table_data, colWidths=[24 * mm, 18 * mm, 80 * mm, 24 * mm, 36 * mm, 30 * mm, 18 * mm, 18 * mm, 18 * mm], repeatRows=1)
    items_table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0B3D91")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, -1), 6.5), ("GRID", (0, 0), (-1, -1), .25, colors.HexColor("#CBD5E1")), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")])]))
    story.append(items_table)
    document.build(story)
    return path
