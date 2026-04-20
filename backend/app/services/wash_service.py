from __future__ import annotations

import calendar
import json
import shutil
import tempfile
from collections import defaultdict
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from app.extensions import db
from app.models import MaintenanceScheduleItem, Vehicle, WashBlockedDay, WashPlanConfig, WashQueueItem, WashRecord, WashScheduleDecision


WASH_FILE_PATTERN = "CONTROLE_DE_LAVAGEM*.xlsx"
WASH_VALUES_FILE = Path(__file__).resolve().parents[3] / "backend" / "data" / "wash_category_values.json"
WASH_CATEGORY_VALUES = {
    "CONJUNTO": Decimal("370.00"),
    "CAVALO": Decimal("240.00"),
    "TERBERG": Decimal("240.00"),
    "CARRETA": Decimal("130.00"),
    "MEDIO PORTE": Decimal("180.00"),
    "PEQUENO PORTE": Decimal("50.00"),
    "CAMINHAO PIPA": Decimal("270.00"),
    "CARRETA PIPA": Decimal("130.00"),
    "ONIBUS": Decimal("200.00"),
    "AMBULANCIA": Decimal("280.00"),
}
SHIFT_ORDER = ("MANHA", "TARDE")
MONTH_NAMES = {
    1: "Janeiro",
    2: "Fevereiro",
    3: "Marco",
    4: "Abril",
    5: "Maio",
    6: "Junho",
    7: "Julho",
    8: "Agosto",
    9: "Setembro",
    10: "Outubro",
    11: "Novembro",
    12: "Dezembro",
}
AUXILIARY_DEFAULTS = [
    {"frota": "MONTANA 1151 CA-03", "placa": "JXQ-1151", "modelo": "CHEVROLET/MONTANA"},
    {"frota": "GOL CA-04", "placa": "OAN-5485", "modelo": "VW GOL"},
    {"frota": "MONTANA SPORT", "placa": "TRX-9H70", "modelo": "CHEVROLET/MONTANA SPORT"},
    {"frota": "CAP-515 BRIGADA 02", "placa": "BRIGADA 02", "modelo": "CAPACITY"},
    {"frota": "ONIBUS CA-01", "placa": "QZK-7E96", "modelo": "ONIBUS"},
    {"frota": "VAN - CA-02", "placa": "NOX-8272", "modelo": "VAN"},
    {"frota": "CAMINHAO PIPA(BRIGADA 01)", "placa": "NON-5189", "modelo": "VOLVO"},
    {"frota": "VW CONSTELLATION CA-06", "placa": "OAK-4703", "modelo": "VW CONSTELLATION"},
    {"frota": "MERCEDES CA-05", "placa": "OAH-8937", "modelo": "MERCEDES"},
    {"frota": "AMBULANCIA USB-1 SPRINTER", "placa": "S/PLACA", "modelo": "MERCEDES"},
    {"frota": "AMBULANCIA USB-2 DUCATO", "placa": "S/PLACA", "modelo": "FIAT"},
]


def _clean(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _norm_key(value: str | None) -> str:
    return (_clean(value) or "").upper()


def _is_terberg_reference(reference: str | None) -> bool:
    normalized = _norm_key(reference)
    if normalized.startswith("TB"):
        return True
    if normalized.isdigit():
        return 2301 <= int(normalized) <= 2310
    return False


def _as_decimal(value) -> Decimal | None:
    if value in {None, ""}:
        return None
    return Decimal(str(value))


def _ordered_category_values(values: dict[str, Decimal] | None = None) -> list[dict]:
    resolved = values or WASH_CATEGORY_VALUES
    return [
        {
            "categoria": category,
            "valor_unitario": float(amount),
        }
        for category, amount in resolved.items()
    ]


def _load_custom_wash_values() -> dict[str, Decimal]:
    if not WASH_VALUES_FILE.exists():
        return {}
    try:
        payload = json.loads(WASH_VALUES_FILE.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}

    values: dict[str, Decimal] = {}
    for key, value in (payload or {}).items():
        amount = _as_decimal(value)
        if key and amount is not None:
            values[_norm_key(key)] = amount
    return values


def save_wash_category_values(values_payload: list[dict]) -> list[dict]:
    values = dict(WASH_CATEGORY_VALUES)
    for item in values_payload:
        category = _norm_key(item.get("categoria"))
        amount = _as_decimal(item.get("valor_unitario"))
        if category and amount is not None:
            values[category] = amount

    WASH_VALUES_FILE.parent.mkdir(parents=True, exist_ok=True)
    serializable = {category: f"{amount:.2f}" for category, amount in values.items()}
    WASH_VALUES_FILE.write_text(json.dumps(serializable, ensure_ascii=False, indent=2), encoding="utf-8")
    return _ordered_category_values(values)


def _vehicle_text(vehicle: Vehicle) -> str:
    return " ".join(
        filter(
            None,
            [
                _norm_key(vehicle.frota),
                _norm_key(vehicle.modelo),
                _norm_key(vehicle.atividade),
            ],
        )
    )


def _infer_auxiliary_category(vehicle: Vehicle) -> str:
    text = _vehicle_text(vehicle)
    frota = _norm_key(vehicle.frota)

    # Regras operacionais fixas para a frota CA.
    if "AMBULANCIA" in text:
        return "AMBULANCIA"
    if "CA-01" in frota:
        return "ONIBUS"
    if "CA-02" in frota:
        return "MEDIO PORTE"
    if "CA-03" in frota or "CA-04" in frota:
        return "PEQUENO PORTE"
    if "CA-05" in frota or "CA-06" in frota:
        return "CAVALO"

    if "ONIBUS" in text:
        return "ONIBUS"
    if "CAMINHAO PIPA" in text:
        return "CAMINHAO PIPA"
    if _is_terberg_reference(vehicle.frota):
        return "TERBERG"
    if any(keyword in text for keyword in ("CONSTELLATION", "MERCEDES", "CAPACITY", "VAN", "SPRINTER", "DUCATO")):
        return "MEDIO PORTE"
    return "PEQUENO PORTE"


def available_wash_modes(vehicle: Vehicle) -> list[str]:
    reference = _norm_key(vehicle.frota)
    if reference.startswith("CV"):
        return ["CAVALO", "CONJUNTO", "CARRETA"]
    if _is_terberg_reference(reference):
        return ["TERBERG", "CONJUNTO", "CARRETA"]
    return [suggest_wash_category(vehicle)]


def wash_mode_requires_trailer(category: str) -> bool:
    return _norm_key(category) in {"CONJUNTO", "CARRETA", "CARRETA PIPA"}


def suggest_wash_category(vehicle: Vehicle, requested_category: str | None = None, carreta: str | None = None) -> str:
    if requested_category:
        return _norm_key(requested_category)

    reference = _norm_key(vehicle.frota)
    if reference.startswith("CV"):
        return "CONJUNTO" if _clean(carreta) else "CAVALO"
    if _is_terberg_reference(reference):
        return "CONJUNTO" if _clean(carreta) else "TERBERG"
    return _infer_auxiliary_category(vehicle)


def suggest_wash_value(category: str) -> Decimal:
    return WASH_CATEGORY_VALUES.get(_norm_key(category), Decimal("0"))


def load_wash_category_values(path: Path | None = None) -> dict[str, Decimal]:
    workbook_path = path or discover_wash_file()
    if not workbook_path:
        values = dict(WASH_CATEGORY_VALUES)
        values.update(_load_custom_wash_values())
        return values

    workbook = _load_workbook_copy(workbook_path)
    if "VALORES" not in workbook.sheetnames:
        values = dict(WASH_CATEGORY_VALUES)
        values.update(_load_custom_wash_values())
        return values

    values = dict(WASH_CATEGORY_VALUES)
    sheet = workbook["VALORES"]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        category = _norm_key(row[0] if len(row) > 0 else None)
        amount = _as_decimal(row[2] if len(row) > 2 else None)
        if category and amount is not None:
            values[category] = amount
    values.update(_load_custom_wash_values())
    return values


def discover_wash_file(explicit_path: str | None = None) -> Path | None:
    if explicit_path:
        path = Path(explicit_path).expanduser()
        if path.exists():
            return path

    home = Path.home()
    for candidate in home.glob(f"OneDrive*/*{WASH_FILE_PATTERN}"):
        if candidate.is_file():
            return candidate

    project_root = Path(__file__).resolve().parents[3]
    for candidate in project_root.glob(WASH_FILE_PATTERN):
        if candidate.is_file():
            return candidate

    return None


def _load_workbook_copy(source_path: Path):
    with tempfile.NamedTemporaryFile(suffix=source_path.suffix, delete=False) as tmp:
        tmp_path = Path(tmp.name)
    try:
        shutil.copy2(source_path, tmp_path)
        return load_workbook(tmp_path, data_only=True)
    finally:
        tmp_path.unlink(missing_ok=True)


def _is_wash_reference_vehicle(vehicle: Vehicle) -> bool:
    return vehicle.ativo and (vehicle.tipo or "").lower() != "carreta"


def _queue_category(vehicle: Vehicle) -> str:
    reference = _norm_key(vehicle.frota)
    return "cavalo" if reference.startswith("CV") or _is_terberg_reference(reference) else "auxiliar"


def ensure_auxiliary_vehicles(path: Path | None = None) -> dict:
    created = 0
    workbook_path = path
    existing = {_norm_key(item.frota) for item in Vehicle.query.all()}

    def create_vehicle(frota: str, placa: str | None, modelo: str):
        nonlocal created
        if frota in existing:
            return
        db.session.add(
            Vehicle(
                frota=frota,
                placa=_norm_key(placa) or "S/PLACA",
                modelo=modelo,
                tipo="auxiliar",
                atividade="VEICULO AUXILIAR",
                status="ON",
                ativo=True,
            )
        )
        existing.add(frota)
        created += 1

    for item in AUXILIARY_DEFAULTS:
        create_vehicle(_norm_key(item["frota"]), item["placa"], item["modelo"])

    should_load_workbook = workbook_path is not None or Vehicle.query.filter(db.func.lower(Vehicle.tipo) == "auxiliar").count() == 0
    if should_load_workbook:
        workbook_path = workbook_path or discover_wash_file()
    if workbook_path:
        workbook = _load_workbook_copy(workbook_path)
        if "FROTA GERAL" in workbook.sheetnames:
            sheet = workbook["FROTA GERAL"]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                frota = _norm_key(row[13] if len(row) > 13 else None)
                if not frota:
                    continue
                placa = _norm_key(row[14] if len(row) > 14 else None)
                modelo = _clean(row[15] if len(row) > 15 else None) or "VEICULO AUXILIAR"
                create_vehicle(frota, placa, modelo)

    if created:
        db.session.commit()
    return {"created": created}


def get_active_trailers() -> list[dict]:
    trailers = (
        Vehicle.query.filter(
            Vehicle.ativo.is_(True),
            db.func.lower(Vehicle.tipo) == "carreta",
        )
        .order_by(Vehicle.frota.asc())
        .all()
    )
    return [
        {
            "id": item.id,
            "frota": item.frota,
            "placa": item.placa,
            "modelo": item.modelo,
        }
        for item in trailers
    ]


def _normalize_queue_positions() -> None:
    items = (
        WashQueueItem.query.join(Vehicle, WashQueueItem.vehicle_id == Vehicle.id)
        .filter(Vehicle.ativo.is_(True), Vehicle.tipo != "carreta")
        .order_by(WashQueueItem.queue_position.asc(), WashQueueItem.id.asc())
        .all()
    )
    for index, item in enumerate(items, start=1):
        item.queue_position = index


def _get_or_create_plan(year: int, month: int) -> WashPlanConfig:
    config = WashPlanConfig.query.filter_by(year=year, month=month).first()
    if config:
        return config
    config = WashPlanConfig(year=year, month=month, morning_capacity=2, afternoon_capacity=2, auxiliary_interval_days=15)
    db.session.add(config)
    db.session.commit()
    return config


def update_plan_config(
    *,
    year: int,
    month: int,
    morning_capacity: int,
    afternoon_capacity: int,
    auxiliary_interval_days: int,
    notes: str | None,
) -> WashPlanConfig:
    config = _get_or_create_plan(year, month)
    config.morning_capacity = max(0, int(morning_capacity))
    config.afternoon_capacity = max(0, int(afternoon_capacity))
    config.auxiliary_interval_days = max(1, int(auxiliary_interval_days))
    config.notes = _clean(notes)
    db.session.commit()
    return config


def set_blocked_day(*, year: int, month: int, day_date: date, shift: str, blocked: bool, reason: str | None) -> WashPlanConfig:
    config = _get_or_create_plan(year, month)
    shift = _norm_key(shift) or "ALL"
    item = WashBlockedDay.query.filter_by(config_id=config.id, day_date=day_date, shift=shift).first()
    if blocked:
        if not item:
            item = WashBlockedDay(config_id=config.id, day_date=day_date, shift=shift, reason=_clean(reason))
            db.session.add(item)
        else:
            item.reason = _clean(reason)
    elif item:
        db.session.delete(item)
    db.session.commit()
    return config


def sync_wash_queue(path: Path | None = None) -> dict:
    ensure_auxiliary_vehicles(path)

    eligible = (
        Vehicle.query.filter(Vehicle.ativo.is_(True), Vehicle.tipo != "carreta")
        .order_by(Vehicle.frota.asc())
        .all()
    )
    queue_items = {item.vehicle_id: item for item in WashQueueItem.query.all()}
    max_position = max([item.queue_position for item in queue_items.values()], default=0)
    created = 0

    for vehicle in eligible:
        item = queue_items.get(vehicle.id)
        if item:
            item.referencia = _norm_key(vehicle.frota)
            item.categoria = _queue_category(vehicle)
            continue
        max_position += 1
        db.session.add(
            WashQueueItem(
                vehicle_id=vehicle.id,
                referencia=_norm_key(vehicle.frota),
                categoria=_queue_category(vehicle),
                queue_position=max_position,
            )
        )
        created += 1

    _normalize_queue_positions()
    db.session.commit()
    imported = import_wash_history(path if WashRecord.query.count() == 0 else None)
    return {"created": created, "history": imported}


def reclassify_wash_queue_categories() -> dict:
    updated = 0
    items = WashQueueItem.query.join(Vehicle, WashQueueItem.vehicle_id == Vehicle.id).all()
    for item in items:
        expected = _queue_category(item.vehicle)
        if item.categoria != expected:
            item.categoria = expected
            updated += 1
    db.session.commit()
    return {"updated": updated}


def import_wash_history(path: Path | None = None) -> dict:
    if WashRecord.query.count() > 0:
        return {"imported": 0, "skipped": True}

    workbook_path = path or discover_wash_file()
    if not workbook_path:
        return {"imported": 0, "skipped": True}

    workbook = _load_workbook_copy(workbook_path)
    if "LAVAGEM" not in workbook.sheetnames:
        return {"imported": 0, "skipped": True}

    imported = 0
    sheet = workbook["LAVAGEM"]
    vehicles_by_ref = {_norm_key(vehicle.frota): vehicle for vehicle in Vehicle.query.all()}
    queue_by_vehicle = {item.vehicle_id: item for item in WashQueueItem.query.all()}
    shifts_by_day = defaultdict(int)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        referencia = _norm_key(row[1] if len(row) > 1 else None)
        if not referencia:
            continue
        vehicle = vehicles_by_ref.get(referencia)
        if not vehicle or not _is_wash_reference_vehicle(vehicle):
            continue

        queue_item = queue_by_vehicle.get(vehicle.id)
        if not queue_item:
            queue_item = WashQueueItem(
                vehicle_id=vehicle.id,
                referencia=referencia,
                categoria=_queue_category(vehicle),
                queue_position=max([item.queue_position for item in queue_by_vehicle.values()], default=0) + 1,
            )
            db.session.add(queue_item)
            db.session.flush()
            queue_by_vehicle[vehicle.id] = queue_item

        wash_dt = row[4] if len(row) > 4 else None
        if isinstance(wash_dt, date) and not isinstance(wash_dt, datetime):
            wash_dt = datetime.combine(wash_dt, time.min)
        if not isinstance(wash_dt, datetime):
            continue

        shift_key = wash_dt.date().isoformat()
        shift_index = shifts_by_day[shift_key]
        shifts_by_day[shift_key] += 1
        turno = "MANHA" if shift_index % 2 == 0 else "TARDE"

        category = _norm_key(row[3] if len(row) > 3 else None) or suggest_wash_category(vehicle, carreta=row[2])
        valor_decimal = _as_decimal(row[6] if len(row) > 6 else None) or suggest_wash_value(category)

        db.session.add(
            WashRecord(
                queue_item_id=queue_item.id,
                vehicle_id=vehicle.id,
                referencia=referencia,
                carreta=_norm_key(row[2] if len(row) > 2 else None),
                tipo_equipamento=category,
                turno=turno,
                status="LAVADO",
                wash_date=wash_dt,
                local=_clean(row[5] if len(row) > 5 else None),
                valor=valor_decimal,
                observacao=None,
                queue_before=None,
                queue_after=None,
            )
        )
        if queue_item.last_wash_at is None or wash_dt > queue_item.last_wash_at:
            queue_item.last_wash_at = wash_dt
            queue_item.last_location = _clean(row[5] if len(row) > 5 else None)
            queue_item.last_value = valor_decimal
        imported += 1

    db.session.flush()
    reorder_queue_by_last_wash()
    db.session.commit()
    return {"imported": imported, "skipped": False}


def reorder_queue_by_last_wash() -> None:
    items = (
        WashQueueItem.query.join(Vehicle, WashQueueItem.vehicle_id == Vehicle.id)
        .filter(Vehicle.ativo.is_(True), Vehicle.tipo != "carreta")
        .order_by(WashQueueItem.last_wash_at.asc().nullsfirst(), WashQueueItem.referencia.asc())
        .all()
    )
    for position, item in enumerate(items, start=1):
        item.queue_position = position


def compute_preventive_date(item: WashQueueItem, reference_day: date | None = None) -> date | None:
    if not item.preventive_enabled or item.preventive_week_of_month is None or item.preventive_weekday is None:
        return None

    current = reference_day or date.today()
    year = current.year
    month = current.month

    for _ in range(2):
        calendar_weeks = calendar.monthcalendar(year, month)
        candidates = [week[item.preventive_weekday] for week in calendar_weeks if week[item.preventive_weekday] != 0]
        if candidates:
            index = min(item.preventive_week_of_month - 1, len(candidates) - 1)
            scheduled = date(year, month, candidates[index])
            if scheduled >= current:
                return scheduled
        month += 1
        if month > 12:
            month = 1
            year += 1
    return None


def apply_preventive_schedule(
    queue_items: list[WashQueueItem],
    *,
    week_of_month: int,
    weekday: int,
    notes: str | None,
) -> int:
    for item in queue_items:
        item.preventive_enabled = True
        item.preventive_week_of_month = week_of_month
        item.preventive_weekday = weekday
        item.preventive_notes = _clean(notes)
    db.session.commit()
    return len(queue_items)


def mark_unavailable(queue_item: WashQueueItem, reason: str | None, user_id: int | None) -> WashQueueItem:
    if queue_item.indisponivel:
        return queue_item
    queue_item.indisponivel = True
    queue_item.motivo_indisponivel = _clean(reason)
    queue_item.indisponivel_desde = datetime.utcnow()
    db.session.add(
        WashRecord(
            queue_item_id=queue_item.id,
            vehicle_id=queue_item.vehicle_id,
            created_by_user_id=user_id,
            referencia=queue_item.referencia,
            carreta=None,
            tipo_equipamento="INDISPONIVEL",
            turno=None,
            status="INDISPONIVEL",
            wash_date=datetime.utcnow(),
            local=None,
            valor=None,
            observacao=queue_item.motivo_indisponivel,
            queue_before=queue_item.queue_position,
            queue_after=queue_item.queue_position,
        )
    )
    db.session.commit()
    return queue_item


def release_unavailable(queue_item: WashQueueItem) -> WashQueueItem:
    queue_item.indisponivel = False
    queue_item.motivo_indisponivel = None
    queue_item.indisponivel_desde = None
    db.session.commit()
    return queue_item


def register_wash(
    queue_item: WashQueueItem,
    *,
    wash_date: datetime,
    location: str | None,
    value,
    carreta: str | None,
    category: str | None,
    shift: str | None,
    notes: str | None,
    photo_path: str | None,
    user_id: int | None,
) -> WashQueueItem:
    vehicle = queue_item.vehicle
    final_category = suggest_wash_category(vehicle, category, carreta)
    final_value = _as_decimal(value) or suggest_wash_value(final_category)
    final_shift = _norm_key(shift) or ("MANHA" if wash_date.hour < 12 else "TARDE")

    if final_category in {"CONJUNTO", "CARRETA", "CARRETA PIPA"} and not _clean(carreta):
        raise ValueError("Informe a carreta quando a lavagem for de conjunto ou somente carreta.")

    _normalize_queue_positions()
    queue_before = queue_item.queue_position
    max_position = db.session.query(db.func.max(WashQueueItem.queue_position)).scalar() or queue_before

    queue_item.last_wash_at = wash_date
    queue_item.last_location = _clean(location)
    queue_item.last_value = final_value
    queue_item.indisponivel = False
    queue_item.motivo_indisponivel = None
    queue_item.indisponivel_desde = None
    queue_item.queue_position = max_position + 1 if max_position >= queue_before else queue_before

    db.session.add(
        WashRecord(
            queue_item_id=queue_item.id,
            vehicle_id=queue_item.vehicle_id,
            created_by_user_id=user_id,
            referencia=queue_item.referencia,
            carreta=_norm_key(carreta),
            tipo_equipamento=final_category,
            turno=final_shift,
            status="LAVADO",
            wash_date=wash_date,
            local=_clean(location),
            valor=final_value,
            observacao=_clean(notes),
            foto_path=_clean(photo_path),
            queue_before=queue_before,
            queue_after=queue_item.queue_position,
        )
    )

    _normalize_queue_positions()
    db.session.commit()
    return queue_item


def mark_schedule_not_completed(
    *,
    queue_item: WashQueueItem,
    scheduled_date: date,
    shift: str,
    reason: str | None,
    user_id: int | None,
) -> WashScheduleDecision:
    normalized_shift = _norm_key(shift) or "MANHA"
    decision = WashScheduleDecision.query.filter_by(
        queue_item_id=queue_item.id,
        scheduled_date=scheduled_date,
        shift=normalized_shift,
    ).first()
    if not decision:
        decision = WashScheduleDecision(
            queue_item_id=queue_item.id,
            vehicle_id=queue_item.vehicle_id,
            decided_by_user_id=user_id,
            scheduled_date=scheduled_date,
            shift=normalized_shift,
            status="NAO_CUMPRIDO",
            reason=_clean(reason),
        )
        db.session.add(decision)
    else:
        decision.decided_by_user_id = user_id
        decision.status = "NAO_CUMPRIDO"
        decision.reason = _clean(reason)
    db.session.commit()
    return decision


def clear_schedule_decision(*, queue_item: WashQueueItem, scheduled_date: date, shift: str) -> bool:
    normalized_shift = _norm_key(shift) or "MANHA"
    decision = WashScheduleDecision.query.filter_by(
        queue_item_id=queue_item.id,
        scheduled_date=scheduled_date,
        shift=normalized_shift,
    ).first()
    if not decision:
        return False
    db.session.delete(decision)
    db.session.commit()
    return True


def reopen_completed_schedule_item(*, queue_item: WashQueueItem, scheduled_date: date, shift: str) -> bool:
    normalized_shift = _norm_key(shift) or "MANHA"
    record = (
        WashRecord.query.filter(
            WashRecord.queue_item_id == queue_item.id,
            WashRecord.status == "LAVADO",
            WashRecord.turno == normalized_shift,
            db.extract("year", WashRecord.wash_date) == scheduled_date.year,
            db.extract("month", WashRecord.wash_date) == scheduled_date.month,
            db.extract("day", WashRecord.wash_date) == scheduled_date.day,
        )
        .order_by(WashRecord.id.desc())
        .first()
    )
    if not record:
        return False

    db.session.delete(record)
    db.session.flush()

    previous_record = (
        WashRecord.query.filter(
            WashRecord.queue_item_id == queue_item.id,
            WashRecord.status == "LAVADO",
        )
        .order_by(WashRecord.wash_date.desc(), WashRecord.id.desc())
        .first()
    )
    if previous_record:
        queue_item.last_wash_at = previous_record.wash_date
        queue_item.last_location = previous_record.local
        queue_item.last_value = previous_record.valor
    else:
        queue_item.last_wash_at = None
        queue_item.last_location = None
        queue_item.last_value = None

    reorder_queue_by_last_wash()
    db.session.commit()
    return True


def _build_blocked_lookup(config: WashPlanConfig) -> dict[tuple[date, str], str | None]:
    blocked = {}
    for item in config.blocked_days:
        blocked[(item.day_date, item.shift)] = item.reason
    return blocked


def _available_slots(config: WashPlanConfig) -> list[dict]:
    blocked = _build_blocked_lookup(config)
    slots = []
    _, last_day = calendar.monthrange(config.year, config.month)
    for day in range(1, last_day + 1):
        current_day = date(config.year, config.month, day)
        for shift, capacity in (("MANHA", config.morning_capacity), ("TARDE", config.afternoon_capacity)):
            if capacity <= 0:
                continue
            if (current_day, "ALL") in blocked or (current_day, shift) in blocked:
                slots.append(
                    {
                        "date": current_day,
                        "shift": shift,
                        "capacity": 0,
                        "blocked": True,
                        "reason": blocked.get((current_day, shift)) or blocked.get((current_day, "ALL")),
                        "items": [],
                    }
                )
                continue
            slots.append(
                {
                    "date": current_day,
                    "shift": shift,
                    "capacity": capacity,
                    "blocked": False,
                    "reason": None,
                    "items": [],
                }
            )
    return slots


def _slot_matches(slot: dict, target_date: date, shift: str) -> bool:
    return slot["date"] == target_date and slot["shift"] == shift and not slot["blocked"]


def _append_to_slot(slots: list[dict], target_date: date, shift: str, item_payload: dict) -> bool:
    for slot in slots:
        if _slot_matches(slot, target_date, shift) and len(slot["items"]) < slot["capacity"]:
            slot["items"].append(item_payload)
            return True
    return False


def _find_first_open_slot(slots: list[dict], start_date: date, preferred_shift: str | None = None) -> dict | None:
    for slot in slots:
        if slot["blocked"] or slot["date"] < start_date:
            continue
        if preferred_shift and slot["shift"] != preferred_shift:
            continue
        if len(slot["items"]) < slot["capacity"]:
            return slot
    if preferred_shift:
        return _find_first_open_slot(slots, start_date, None)
    return None


def _find_previous_open_slot(slots: list[dict], target_date: date, preferred_shift: str | None = None) -> dict | None:
    for slot in reversed(slots):
        if slot["blocked"] or slot["date"] > target_date:
            continue
        if preferred_shift and slot["shift"] != preferred_shift:
            continue
        if len(slot["items"]) < slot["capacity"]:
            return slot
    if preferred_shift:
        return _find_previous_open_slot(slots, target_date, None)
    return None


def _payload_for_queue_item(item: WashQueueItem, category: str | None = None) -> dict:
    vehicle = item.vehicle
    final_category = category or suggest_wash_category(vehicle)
    return {
        "queue_item_id": item.id,
        "referencia": item.referencia,
        "placa": vehicle.placa if vehicle else "-",
        "modelo": vehicle.modelo if vehicle else "-",
        "categoria_lavagem": final_category,
        "valor_sugerido": float(suggest_wash_value(final_category)),
        "tipo": item.categoria,
        "foto_path": None,
    }


def _append_maintenance_preventive_washes(slots: list[dict], *, year: int, month: int, today: date) -> None:
    first_day = date(year, month, 1)
    _, last_number = calendar.monthrange(year, month)
    last_day = date(year, month, last_number)
    preventive_items = (
        MaintenanceScheduleItem.query.join(MaintenanceScheduleItem.schedule)
        .filter(
            MaintenanceScheduleItem.scheduled_date.isnot(None),
            MaintenanceScheduleItem.scheduled_date >= first_day + timedelta(days=1),
            MaintenanceScheduleItem.scheduled_date <= last_day + timedelta(days=1),
            MaintenanceScheduleItem.status.notin_(["CANCELADO"]),
            MaintenanceScheduleItem.schedule.has(source_type="PREVENTIVA"),
        )
        .order_by(MaintenanceScheduleItem.scheduled_date.asc(), MaintenanceScheduleItem.id.asc())
        .all()
    )

    planned_keys = {
        (item.get("queue_item_id"), slot["date"])
        for slot in slots
        for item in slot["items"]
        if item.get("queue_item_id")
    }
    queue_by_vehicle = {item.vehicle_id: item for item in WashQueueItem.query.all()}
    for preventive_item in preventive_items:
        wash_date = preventive_item.scheduled_date - timedelta(days=1)
        if wash_date < first_day or wash_date > last_day or wash_date < today:
            continue
        queue_item = queue_by_vehicle.get(preventive_item.vehicle_id)
        if not queue_item or (queue_item.id, wash_date) in planned_keys:
            continue
        payload = _payload_for_queue_item(queue_item, "CAVALO")
        payload["origem"] = "PREVENTIVA_MANUTENCAO"
        payload["status_origem"] = "LAVAGEM PRÉ-PREVENTIVA"
        payload["observacao"] = f"Lavagem pré-preventiva para {preventive_item.schedule.title if preventive_item.schedule else 'manutenção preventiva'}"
        target = _find_previous_open_slot(slots, wash_date, "MANHA")
        if target:
            target["items"].append(payload)
            planned_keys.add((queue_item.id, target["date"]))


def _build_month_schedule(year: int, month: int, config: WashPlanConfig, queue_items: list[WashQueueItem], today: date) -> dict:
    slots = _available_slots(config)
    records_by_key = {}

    month_records = (
        WashRecord.query.filter(
            db.extract("year", WashRecord.wash_date) == year,
            db.extract("month", WashRecord.wash_date) == month,
            WashRecord.status == "LAVADO",
        )
        .order_by(WashRecord.wash_date.asc(), WashRecord.turno.asc(), WashRecord.id.asc())
        .all()
    )

    if (year, month) <= (today.year, today.month):
        for record in month_records:
            record_date = record.wash_date.date()
            shift = record.turno or "MANHA"
            records_by_key[(record.queue_item_id, record_date, shift)] = record
            payload = {
                "queue_item_id": record.queue_item_id,
                "referencia": record.referencia,
                "placa": record.vehicle.placa if record.vehicle else "-",
                "modelo": record.vehicle.modelo if record.vehicle else "-",
                "categoria_lavagem": record.tipo_equipamento,
                "valor_sugerido": float(record.valor) if record.valor is not None else float(suggest_wash_value(record.tipo_equipamento)),
                "tipo": record.queue_item.categoria if record.queue_item else "cavalo",
                "carreta": record.carreta,
                "foto_path": record.foto_path,
                "realizado": True,
            }
            if not _append_to_slot(slots, record_date, shift, payload):
                slot = _find_first_open_slot(slots, record_date, shift) or _find_first_open_slot(slots, record_date)
                if slot:
                    slot["items"].append(payload)

    if (year, month) >= (today.year, today.month):
        # When planning the current month, never schedule items into past days.
        # Past days remain for history/visualization, even if they have pending items.
        start_day = date(year, month, 1)
        if (year, month) == (today.year, today.month):
            start_day = today

        planned_refs = {item["referencia"] for slot in slots for item in slot["items"]}
        active_queue = [item for item in queue_items if not item.indisponivel and item.referencia not in planned_refs]
        cavalo_queue = [item for item in active_queue if item.categoria == "cavalo"]
        auxiliary_queue = [item for item in active_queue if item.categoria == "auxiliar"]

        _append_maintenance_preventive_washes(slots, year=year, month=month, today=today)
        planned_refs = {item["referencia"] for slot in slots for item in slot["items"]}
        cavalo_queue = [item for item in cavalo_queue if item.referencia not in planned_refs]
        auxiliary_queue = [item for item in auxiliary_queue if item.referencia not in planned_refs]

        # Preventivas first.
        for item in cavalo_queue[:]:
            preventive_date = compute_preventive_date(item, start_day)
            if preventive_date and preventive_date.year == year and preventive_date.month == month:
                target_shift = "MANHA"
                if _append_to_slot(slots, preventive_date, target_shift, _payload_for_queue_item(item, "CAVALO")):
                    cavalo_queue.remove(item)

        # Auxiliares every configured interval days.
        aux_day = start_day
        aux_index = 0
        while auxiliary_queue and aux_day.month == month:
            slot = _find_first_open_slot(slots, aux_day, "TARDE") or _find_first_open_slot(slots, aux_day)
            if slot:
                item = auxiliary_queue.pop(0)
                slot["items"].append(_payload_for_queue_item(item))
                aux_index += 1
            aux_day += timedelta(days=config.auxiliary_interval_days)

        # Fill remaining slots by queue order.
        all_remaining = cavalo_queue + auxiliary_queue
        for item in all_remaining:
            slot = _find_first_open_slot(slots, start_day)
            if not slot:
                break
            slot["items"].append(_payload_for_queue_item(item))

    decisions = (
        WashScheduleDecision.query.filter(
            db.extract("year", WashScheduleDecision.scheduled_date) == year,
            db.extract("month", WashScheduleDecision.scheduled_date) == month,
        ).all()
    )
    decisions_by_key = {
        (item.queue_item_id, item.scheduled_date, item.shift): item
        for item in decisions
    }

    for slot in slots:
        for item in slot["items"]:
            item["scheduled_date"] = slot["date"].isoformat()
            item["scheduled_shift"] = slot["shift"]
            record = records_by_key.get((item.get("queue_item_id"), slot["date"], slot["shift"]))
            decision = decisions_by_key.get((item.get("queue_item_id"), slot["date"], slot["shift"]))
            if record or item.get("realizado"):
                item["status_execucao"] = "LAVADO"
                item["status_rotulo"] = "Lavado"
                item["status_motivo"] = None
            elif decision:
                item["status_execucao"] = decision.status
                item["status_rotulo"] = "Não lavado"
                item["status_motivo"] = decision.reason
            else:
                item["status_execucao"] = "PENDENTE"
                item["status_rotulo"] = "Pendente"
                item["status_motivo"] = None

    by_day: dict[str, dict] = {}
    for slot in slots:
        key = slot["date"].isoformat()
        day_bucket = by_day.setdefault(
            key,
            {
                "date": key,
                "weekday": slot["date"].weekday(),
                "day": slot["date"].day,
                "morning": [],
                "afternoon": [],
                "blocked": False,
                "blocked_morning": False,
                "blocked_afternoon": False,
                "reason": None,
            },
        )
        if slot["shift"] == "MANHA":
            day_bucket["morning"] = slot["items"]
            day_bucket["blocked_morning"] = slot["blocked"]
        else:
            day_bucket["afternoon"] = slot["items"]
            day_bucket["blocked_afternoon"] = slot["blocked"]
        if slot["blocked"]:
            day_bucket["blocked"] = True
            day_bucket["reason"] = slot["reason"]

    return {
        "config": config.to_dict(),
        "days": sorted(by_day.values(), key=lambda item: item["date"]),
    }


def build_wash_overview(
    reference_day: date | None = None,
    *,
    year: int | None = None,
    month: int | None = None,
) -> dict:
    current = reference_day or date.today()
    target_year = year or current.year
    target_month = month or current.month

    # A visão geral é chamada com frequência pelo desktop. Sincronizar a fila aqui
    # (incluindo leitura de Excel) deixa a UI travada por muito tempo.
    # Mantemos apenas um auto-sync inicial, quando ainda não existe fila.
    if WashQueueItem.query.count() == 0:
        try:
            sync_wash_queue(discover_wash_file())
        except Exception:
            db.session.rollback()

    config = _get_or_create_plan(target_year, target_month)
    category_values = load_wash_category_values()

    queue = (
        WashQueueItem.query.join(Vehicle, WashQueueItem.vehicle_id == Vehicle.id)
        .filter(Vehicle.ativo.is_(True), Vehicle.tipo != "carreta")
        .order_by(WashQueueItem.queue_position.asc())
        .all()
    )

    history = (
        WashRecord.query.filter(
            db.extract("year", WashRecord.wash_date) == target_year,
            db.extract("month", WashRecord.wash_date) == target_month,
        )
        .order_by(WashRecord.wash_date.desc(), WashRecord.id.desc())
        .all()
    )

    queue_payload = []
    preventive_payload = []
    next_available = None
    unavailable_count = 0
    for item in queue:
        next_preventive = compute_preventive_date(item, current)
        if item.indisponivel:
            unavailable_count += 1
        elif next_available is None:
            next_available = item

        data = item.to_dict()
        data["proxima_preventiva"] = next_preventive.isoformat() if next_preventive else None
        data["status_fila"] = "INDISPONIVEL" if item.indisponivel else "DISPONIVEL"
        data["modos_lavagem"] = available_wash_modes(item.vehicle)
        data["categoria_sugerida"] = suggest_wash_category(item.vehicle)
        data["valor_sugerido"] = float(category_values.get(data["categoria_sugerida"], suggest_wash_value(data["categoria_sugerida"])))
        data["carreta_obrigatoria"] = wash_mode_requires_trailer(data["categoria_sugerida"])
        queue_payload.append(data)
        if next_preventive:
            preventive_payload.append(data)

    month_schedule = _build_month_schedule(target_year, target_month, config, queue, current)

    total_value = sum((record.valor or Decimal("0")) for record in history if record.status == "LAVADO")
    by_category = defaultdict(lambda: {"quantidade": 0, "valor": Decimal("0")})
    by_vehicle = defaultdict(lambda: {"quantidade": 0, "valor": Decimal("0")})
    for record in history:
        if record.status != "LAVADO":
            continue
        category_bucket = by_category[record.tipo_equipamento]
        category_bucket["quantidade"] += 1
        category_bucket["valor"] += record.valor or Decimal("0")
        vehicle_bucket = by_vehicle[record.referencia]
        vehicle_bucket["quantidade"] += 1
        vehicle_bucket["valor"] += record.valor or Decimal("0")

    return {
        "periodo": {
            "ano": target_year,
            "mes": target_month,
            "rotulo": f"{MONTH_NAMES.get(target_month, str(target_month))} {target_year}",
        },
        "resumo": {
            "proximo": next_available.to_dict() if next_available else None,
            "lavados_mes": len([record for record in history if record.status == "LAVADO"]),
            "indisponiveis": unavailable_count,
            "programados_preventiva": len(preventive_payload),
            "valor_total": float(total_value),
        },
        "fila": queue_payload,
        "historico": [record.to_dict() for record in history],
        "preventivas": preventive_payload,
        "cronograma": month_schedule,
        "carretas": get_active_trailers(),
        "tabela_valores": _ordered_category_values(category_values),
        "indicadores": {
            "por_categoria": [
                {
                    "categoria": key,
                    "quantidade": value["quantidade"],
                    "valor": float(value["valor"]),
                }
                for key, value in sorted(by_category.items(), key=lambda item: (-item[1]["quantidade"], item[0]))
            ],
            "por_veiculo": [
                {
                    "referencia": key,
                    "quantidade": value["quantidade"],
                    "valor": float(value["valor"]),
                }
                for key, value in sorted(by_vehicle.items(), key=lambda item: (-item[1]["quantidade"], item[0]))
            ],
        },
    }


def build_tomorrow_message_payload(reference_day: date | None = None) -> dict:
    current = reference_day or date.today()
    tomorrow = current + timedelta(days=1)
    overview = build_wash_overview(current, year=tomorrow.year, month=tomorrow.month)
    target = next((item for item in overview["cronograma"]["days"] if item["date"] == tomorrow.isoformat()), None)
    return {
        "date": tomorrow.isoformat(),
        "morning": target["morning"] if target else [],
        "afternoon": target["afternoon"] if target else [],
    }
