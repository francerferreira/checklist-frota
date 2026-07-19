from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook

from app.extensions import db
from app.models import EquipmentFamily, EquipmentProfile, Vehicle
from app.services.availability_service import seed_operational_states
from app.services.equipment_structure_service import seed_equipment_structure
from app.utils.timezone import now_manaus_naive


PORTUARY_TYPES = {
    "LBS": "lbs",
    "RTG": "rtg",
    "SPREADER": "spreader",
}
PORTUARY_REQUIRED_COLUMNS = {"Tipo", "Equipamento", "Modelo", "Numero_Serie", "Ano", "Status"}


def _normalize_text(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _normalize_year(value) -> str | None:
    text = _normalize_text(value)
    if not text:
        return None
    return text


def discover_inventory_file(configured_path: str | None = None) -> Path | None:
    if not configured_path:
        return None

    project_root = Path(__file__).resolve().parents[3]
    path = Path(configured_path).expanduser().resolve()
    if not path.exists() or not path.is_file():
        return None
    if not path.is_relative_to(project_root):
        return None
    return path


def _map_carreta(row: tuple) -> dict:
    status = _normalize_text(row[9]) or "ON"
    return {
        "frota": _normalize_text(row[1]),
        "tipo": "carreta",
        "placa": _normalize_text(row[3]) or "S/PLACA",
        "ano": _normalize_year(row[4]),
        "chassi": _normalize_text(row[5]),
        "configuracao": _normalize_text(row[6]),
        "modelo": _normalize_text(row[7]) or "CARRETA",
        "atividade": _normalize_text(row[8]),
        "status": status,
        "descricao": _normalize_text(row[11]),
        "local": None,
        "ativo": status.upper() != "OFF",
    }


def _map_cavalo(row: tuple) -> dict:
    status = _normalize_text(row[6]) or "ON"
    return {
        "frota": _normalize_text(row[1]),
        "tipo": "cavalo",
        "placa": _normalize_text(row[3]) or "S/PLACA",
        "ano": _normalize_year(row[2]),
        "chassi": _normalize_text(row[5]),
        "configuracao": None,
        "modelo": _normalize_text(row[4]) or "CAVALO MECANICO",
        "atividade": _normalize_text(row[8]),
        "status": status,
        "descricao": _normalize_text(row[8]),
        "local": _normalize_text(row[7]),
        "ativo": status.upper() != "OFF",
    }


def import_inventory_data(path: Path) -> dict:
    workbook = load_workbook(path, read_only=True, data_only=True)
    imported = 0
    updated = 0

    sheets = {
        "CARRETAS": _map_carreta,
        "CAVALOS": _map_cavalo,
    }

    for sheet_name, mapper in sheets.items():
        worksheet = workbook[sheet_name]
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            payload = mapper(row)
            if not payload["frota"]:
                continue

            vehicle = Vehicle.query.filter_by(frota=payload["frota"]).first()
            if vehicle is None:
                vehicle = Vehicle(**payload)
                db.session.add(vehicle)
                imported += 1
                continue

            preserved_photo = vehicle.foto_path
            for key, value in payload.items():
                setattr(vehicle, key, value)
            vehicle.foto_path = preserved_photo
            vehicle.ativo = (vehicle.status or "").upper() not in {"RETIRADO", "OFF"}
            updated += 1

    db.session.commit()
    from app.services.equipment_structure_service import seed_equipment_structure

    seed_equipment_structure()
    workbook.close()
    return {
        "arquivo": str(path),
        "importados": imported,
        "atualizados": updated,
    }


def _normalize_portuary_status(value: str | None) -> tuple[str, bool]:
    status = (_normalize_text(value) or "").strip().upper()
    if status in {"ATIVO", "ON", "DISPONIVEL"}:
        return "ON", True
    if status in {"INATIVO", "OFF", "RETIRADO"}:
        return "OFF", False
    raise ValueError(f"Status portuario invalido: {value!r}.")


def read_portuary_csv(path: Path) -> list[dict]:
    """Read and validate the three-module portuary inventory CSV."""
    rows: list[dict] = []
    seen_frotas: set[str] = set()
    seen_serials: set[str] = set()

    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        columns = set(reader.fieldnames or ())
        missing = PORTUARY_REQUIRED_COLUMNS - columns
        if missing:
            raise ValueError(f"Colunas obrigatorias ausentes: {', '.join(sorted(missing))}.")

        for line_number, raw in enumerate(reader, start=2):
            equipment_type = (_normalize_text(raw.get("Tipo")) or "").upper()
            tipo = PORTUARY_TYPES.get(equipment_type)
            frota = (_normalize_text(raw.get("Equipamento")) or "").upper()
            if not tipo:
                raise ValueError(f"Tipo invalido na linha {line_number}: {equipment_type!r}.")
            if not frota:
                raise ValueError(f"Equipamento vazio na linha {line_number}.")
            if frota in seen_frotas:
                raise ValueError(f"Equipamento duplicado no CSV: {frota}.")
            seen_frotas.add(frota)

            serial = _normalize_text(raw.get("Numero_Serie"))
            if serial:
                serial = serial.upper()
                if serial in seen_serials:
                    raise ValueError(f"Numero de serie duplicado no CSV: {serial}.")
                seen_serials.add(serial)

            status, active = _normalize_portuary_status(raw.get("Status"))
            model = _normalize_text(raw.get("Modelo")) or equipment_type
            rows.append(
                {
                    "frota": frota,
                    "tipo": tipo,
                    "placa": "S/PLACA",
                    "modelo": model.upper(),
                    "ano": _normalize_year(raw.get("Ano")),
                    "status": status,
                    "ativo": active,
                    "serial_number": serial,
                    "capacity": _normalize_text(raw.get("Modelo")),
                    "descricao": f"Inventario portuario - {equipment_type}",
                }
            )

    if not rows:
        raise ValueError("O CSV portuario nao possui registros.")
    return rows


def replace_portuary_inventory(path: Path) -> dict:
    """Replace active fleet visibility while preserving historical records."""
    rows = read_portuary_csv(path)
    seed_equipment_structure()
    families = {
        family.code: family
        for family in EquipmentFamily.query.filter(EquipmentFamily.code.in_(PORTUARY_TYPES.values())).all()
    }
    missing_families = set(PORTUARY_TYPES.values()) - set(families)
    if missing_families:
        raise ValueError(f"Familias portuarias ausentes: {', '.join(sorted(missing_families))}.")

    existing_by_frota = {vehicle.frota.upper(): vehicle for vehicle in Vehicle.query.all()}
    requested_frotas = {row["frota"] for row in rows}
    imported = 0
    updated = 0
    retired = 0

    for row in rows:
        vehicle = existing_by_frota.get(row["frota"])
        if vehicle is None:
            vehicle = Vehicle(
                frota=row["frota"],
                tipo=row["tipo"],
                placa=row["placa"],
                modelo=row["modelo"],
                ano=row["ano"],
                status=row["status"],
                atividade=row["tipo"].upper(),
                descricao=row["descricao"],
                ativo=row["ativo"],
            )
            db.session.add(vehicle)
            imported += 1
        else:
            vehicle.tipo = row["tipo"]
            vehicle.placa = row["placa"]
            vehicle.modelo = row["modelo"]
            vehicle.ano = row["ano"]
            vehicle.status = row["status"]
            vehicle.atividade = row["tipo"].upper()
            vehicle.descricao = row["descricao"]
            vehicle.ativo = row["ativo"]
            vehicle.retirado_em = None if row["ativo"] else vehicle.retirado_em
            updated += 1

        db.session.flush()
        profile = vehicle.equipment_profile
        if profile is None:
            profile = EquipmentProfile(vehicle_id=vehicle.id, family_id=families[row["tipo"]].id)
            db.session.add(profile)
        profile.family_id = families[row["tipo"]].id
        profile.serial_number = row["serial_number"]
        profile.capacity = row["capacity"]
        profile.criticality = "MEDIA"

    for vehicle in existing_by_frota.values():
        if vehicle.ativo and vehicle.frota.upper() not in requested_frotas:
            vehicle.ativo = False
            vehicle.status = "RETIRADO"
            vehicle.retirado_em = now_manaus_naive()
            retired += 1

    db.session.flush()
    seed_operational_states()
    db.session.commit()
    return {
        "arquivo": str(path),
        "total_csv": len(rows),
        "por_modulo": {
            module.upper(): sum(1 for row in rows if row["tipo"] == module)
            for module in PORTUARY_TYPES.values()
        },
        "importados": imported,
        "atualizados": updated,
        "retirados_sem_apagar_historico": retired,
        "cadastro_fisico_excluido": 0,
    }
