from __future__ import annotations

import hashlib
import json
import re
import unicodedata
from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

from app.extensions import db
from app.models import (
    InvoicePurchaseOrderLink,
    Material,
    PurchaseImportBatch,
    PurchaseImportSourceRow,
    PurchaseInvoice,
    PurchaseInvoiceItem,
    PurchaseOrder,
    PurchaseOrderItem,
    PurchaseProcessEvent,
    PurchaseRequest,
    PurchaseRequestItem,
    PurchaseServiceCatalog,
    Supplier,
)
from app.utils.timezone import now_manaus_naive


SOURCE_COLUMNS = (
    "CODCOLIGADA", "CODFILIAL", "DT_EMISSAO_SC", "NUM_SC", "SOLICITANTE", "OBSERVACAO_SC",
    "COD_PROD", "PRODUTO", "MARCA", "REF_MANUAL", "NUM_FABRICANTE", "QUANTIDADE_SC",
    "ITM_FATURADOS", "CODTMV_PC", "COMPRADOR", "NUMERO_PC", "DT_EMISSAO_PC", "QUANTIDADE_PC",
    "CODTMV", "FORN", "SERIE", "NUM_NOTA", "DT_EMISSAO_NF", "ITM_RECEBIDOS", "QUANTIDADE_RECEBIDA", "VLR_ITM",
)
SERVICE_PREFIX = "017"
MODULES = ("LBS", "RTG", "SPREADER", "FROTA")


def _text(value) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\xa0", " ").strip()
    return text or None


def _business_text(value) -> str | None:
    text = _text(value)
    if not text or text == "-" or text.lower().endswith(" total"):
        return None
    return text


def _date(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
    text = str(value).strip()
    for pattern in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    return None


def _decimal(value) -> Decimal:
    if value in (None, "", "-"):
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip().replace(".", "").replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def _json_value(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def _normalize_key(value) -> str:
    text = _text(value) or ""
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", text).strip().upper()


def _module_from_text(*values) -> str:
    text = _normalize_key(" ".join(_text(value) or "" for value in values))
    for module in MODULES:
        if re.search(rf"\b{module}\b", text):
            return module
    return "OUTROS"


def _item_type(product_code: str | None) -> str:
    return "SERVICO" if (_text(product_code) or "").split(".", 1)[0] == SERVICE_PREFIX else "MATERIAL"


def _status_for_rows(rows: list[dict]) -> str:
    total = sum((row["quantity_requested"] for row in rows), Decimal("0"))
    received = sum((row["quantity_received"] for row in rows), Decimal("0"))
    has_pc = any(row["pc_number"] for row in rows)
    has_nf = any(row["invoice_number"] for row in rows)
    if not has_pc:
        return "AGUARDANDO_PC"
    if not has_nf:
        return "AGUARDANDO_NF"
    if received <= 0:
        return "EM_TRANSITO"
    if received < total:
        return "PARCIALMENTE_RECEBIDA"
    return "RECEBIDA"


def _legacy_request_status(status: str) -> str:
    # The compatibility table predates the canonical status engine.
    return {"AGUARDANDO_PC": "SOLICITADA", "AGUARDANDO_NF": "EM_TRANSITO"}.get(status, status)


def _supplier_code(name: str) -> str:
    digest = hashlib.sha1(_normalize_key(name).encode("utf-8")).hexdigest()[:12]
    return f"HIST-{digest}"[:40]


def _material_reference(code: str | None, description: str | None) -> str:
    raw = _text(code) or f"HIST-{hashlib.sha1(_normalize_key(description).encode('utf-8')).hexdigest()[:20]}"
    return raw[:80]


def _read_rows(path: Path) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        headers = tuple(next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True)))
        if headers != SOURCE_COLUMNS:
            missing = sorted(set(SOURCE_COLUMNS) - set(headers))
            extra = sorted(set(headers) - set(SOURCE_COLUMNS))
            raise ValueError(f"Cabeçalho histórico inválido. Ausentes: {missing}; extras: {extra}.")
        rows = []
        for source_row_number, raw in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            payload = {key: raw[index] if index < len(raw) else None for index, key in enumerate(headers)}
            if not _business_text(payload.get("NUM_SC")) or not _business_text(payload.get("COD_PROD")):
                continue
            rows.append({
                "source_row_number": source_row_number,
                "payload": {key: _json_value(value) for key, value in payload.items()},
                "company_code": _business_text(payload.get("CODCOLIGADA")),
                "branch_code": _business_text(payload.get("CODFILIAL")),
                "sc_number": _business_text(payload.get("NUM_SC")),
                "sc_date": _date(payload.get("DT_EMISSAO_SC")),
                "requester_raw": _business_text(payload.get("SOLICITANTE")),
                "justification": _business_text(payload.get("OBSERVACAO_SC")),
                "product_code": _business_text(payload.get("COD_PROD")),
                "description": _business_text(payload.get("PRODUTO")) or "ITEM HISTÓRICO",
                "brand": _business_text(payload.get("MARCA")),
                "manual_reference": _business_text(payload.get("REF_MANUAL")),
                "manufacturer_part_number": _business_text(payload.get("NUM_FABRICANTE")),
                "quantity_requested": _decimal(payload.get("QUANTIDADE_SC")),
                "quantity_ordered": _decimal(payload.get("QUANTIDADE_PC")),
                "quantity_received": _decimal(payload.get("QUANTIDADE_RECEBIDA")),
                "buyer_raw": _business_text(payload.get("COMPRADOR")),
                "pc_number": _business_text(payload.get("NUMERO_PC")),
                "pc_date": _date(payload.get("DT_EMISSAO_PC")),
                "supplier_raw": _business_text(payload.get("FORN")),
                "series": _business_text(payload.get("SERIE")),
                "invoice_number": _business_text(payload.get("NUM_NOTA")),
                "invoice_date": _date(payload.get("DT_EMISSAO_NF")),
                "line_value": _decimal(payload.get("VLR_ITM")),
            })
        return rows
    finally:
        workbook.close()


def import_purchase_workbook(path: str | Path, *, user_id: int) -> dict:
    source_path = Path(path).expanduser().resolve()
    if not source_path.exists() or not source_path.is_file():
        raise FileNotFoundError(f"Arquivo histórico não encontrado: {source_path}")
    checksum = hashlib.sha256(source_path.read_bytes()).hexdigest()
    existing = PurchaseImportBatch.query.filter_by(source_checksum=checksum).first()
    if existing:
        return {"status": "JA_IMPORTADO", "batch": existing.to_dict()}

    rows = _read_rows(source_path)
    batch = PurchaseImportBatch(source_filename=source_path.name, source_checksum=checksum, imported_by_user_id=user_id, rows_read=len(rows), status="PROCESSANDO")
    db.session.add(batch)
    db.session.flush()

    grouped: dict[tuple[str | None, str | None, str], list[dict]] = defaultdict(list)
    for row in rows:
        grouped[(row["company_code"], row["branch_code"], row["sc_number"])].append(row)

    material_cache: dict[str, Material] = {}
    service_cache: dict[str, PurchaseServiceCatalog] = {}
    supplier_cache: dict[str, Supplier] = {}
    request_cache: dict[tuple[str | None, str | None, str], PurchaseRequest] = {}
    item_cache: dict[tuple[int, int], PurchaseRequestItem] = {}
    order_cache: dict[tuple[str, str | None, str | None], PurchaseOrder] = {}
    order_item_cache: dict[tuple[int, int], PurchaseOrderItem] = {}
    invoice_cache: dict[tuple[str | None, str | None, str], PurchaseInvoice] = {}
    created = {"requests": 0, "items": 0, "orders": 0, "order_items": 0, "invoices": 0, "invoice_links": 0, "materials": 0, "services": 0, "suppliers": 0}

    for request_key, request_rows in grouped.items():
        company_code, branch_code, sc_number = request_key
        request = PurchaseRequest.query.filter_by(company_code=company_code, branch_code=branch_code, sc_number=sc_number).first()
        if request is None:
            request = PurchaseRequest(
                code=f"SC-{company_code or '0'}-{branch_code or '0'}-{sc_number}"[:40], material_id=None,
                requested_quantity=max(1, int(sum((row["quantity_requested"] for row in request_rows), Decimal("0")))),
                received_quantity=min(max(0, int(sum((row["quantity_received"] for row in request_rows), Decimal("0")))), max(1, int(sum((row["quantity_requested"] for row in request_rows), Decimal("0"))))),
                status=_legacy_request_status(_status_for_rows(request_rows)), priority="MEDIA", created_by_user_id=user_id,
                company_code=company_code, branch_code=branch_code, sc_number=sc_number,
                sc_date=request_rows[0]["sc_date"], requester_raw=request_rows[0]["requester_raw"],
                request_type=None, module=_module_from_text(request_rows[0]["justification"], request_rows[0]["description"]),
                equipment_raw=None, justification=request_rows[0]["justification"], imported=True, import_batch_id=batch.id,
            )
            db.session.add(request)
            db.session.flush()
            created["requests"] += 1
        else:
            request.status = _legacy_request_status(_status_for_rows(request_rows))
            request.imported = True
            request.import_batch_id = batch.id
        request_cache[request_key] = request
        material_count = service_count = 0

        for line_number, row in enumerate(request_rows, start=1):
            item_key = (request.id, line_number)
            item = item_cache.get(item_key) or PurchaseRequestItem.query.filter_by(purchase_request_id=request.id, line_number=line_number).first()
            item_type = _item_type(row["product_code"])
            material = service = None
            if item_type == "MATERIAL":
                reference = _material_reference(row["product_code"], row["description"])
                material = material_cache.get(reference) or Material.query.filter_by(referencia=reference).first()
                if material is None:
                    material = Material(referencia=reference, descricao=row["description"][:255], aplicacao_tipo="ambos", quantidade_estoque=0)
                    db.session.add(material)
                    db.session.flush()
                    created["materials"] += 1
                material_cache[reference] = material
                material_count += 1
            else:
                code = (row["product_code"] or _material_reference(None, row["description"]))[:80]
                service = service_cache.get(code) or PurchaseServiceCatalog.query.filter_by(code=code).first()
                if service is None:
                    service = PurchaseServiceCatalog(code=code, service_name=row["description"][:255], description=row["description"], active=True, imported=True)
                    db.session.add(service)
                    db.session.flush()
                    created["services"] += 1
                service_cache[code] = service
                service_count += 1

            if item is None:
                item = PurchaseRequestItem(purchase_request_id=request.id, line_number=line_number, description_raw=row["description"], quantity_requested=row["quantity_requested"] or Decimal("1"), imported=True)
                db.session.add(item)
                created["items"] += 1
            item.item_type = item_type
            item.material_id = material.id if material else None
            item.service_catalog_id = service.id if service else None
            item.product_code_raw = row["product_code"]
            item.description_raw = row["description"]
            item.brand_raw = row["brand"]
            item.manual_reference_raw = row["manual_reference"]
            item.manufacturer_part_number_raw = row["manufacturer_part_number"]
            item.quantity_requested = row["quantity_requested"] or Decimal("1")
            item.quantity_received = min(row["quantity_received"], item.quantity_requested)
            item.status = "RECEBIDO" if item.quantity_received >= item.quantity_requested else ("AGUARDANDO_PC" if not row["pc_number"] else ("AGUARDANDO_NF" if not row["invoice_number"] else "RECEBIMENTO_PARCIAL"))
            db.session.flush()
            item_cache[item_key] = item

            if row["supplier_raw"]:
                key = _normalize_key(row["supplier_raw"])
                supplier = supplier_cache.get(key) or Supplier.query.filter_by(name=row["supplier_raw"][:180]).first()
                if supplier is None:
                    supplier = Supplier(code=_supplier_code(row["supplier_raw"]), name=row["supplier_raw"][:180], legal_name=row["supplier_raw"][:220], active=True)
                    db.session.add(supplier)
                    db.session.flush()
                    created["suppliers"] += 1
                supplier_cache[key] = supplier
            else:
                supplier = None

            if row["pc_number"]:
                order_key = (row["pc_number"], company_code, branch_code)
                order = order_cache.get(order_key) or PurchaseOrder.query.filter_by(pc_number=row["pc_number"], company_code=company_code, branch_code=branch_code).first()
                if order is None:
                    order = PurchaseOrder(pc_number=row["pc_number"], pc_date=row["pc_date"], supplier_id=supplier.id if supplier else None, supplier_raw=row["supplier_raw"], total_value=Decimal("0"), status="RECEBIDO" if row["invoice_number"] else "EMITIDO", imported=True, import_batch_id=batch.id)
                    db.session.add(order)
                    db.session.flush()
                    created["orders"] += 1
                order.supplier_id = supplier.id if supplier else order.supplier_id
                order.supplier_raw = row["supplier_raw"] or order.supplier_raw
                order.total_value = (order.total_value or Decimal("0")) + row["line_value"]
                order_cache[order_key] = order
                order_item_key = (order.id, item.id)
                order_item = order_item_cache.get(order_item_key) or PurchaseOrderItem.query.filter_by(purchase_order_id=order.id, purchase_request_item_id=item.id).first()
                if order_item is None:
                    order_item = PurchaseOrderItem(purchase_order_id=order.id, purchase_request_item_id=item.id, quantity_ordered=row["quantity_ordered"] or row["quantity_requested"] or Decimal("1"), total_price=row["line_value"], status="RECEBIDO" if row["invoice_number"] else "EMITIDO")
                    db.session.add(order_item)
                    db.session.flush()
                    created["order_items"] += 1
                order_item_cache[order_item_key] = order_item
            else:
                order = order_item = None

            if row["invoice_number"]:
                invoice_key = (row["invoice_number"], row["series"], supplier.id if supplier else None)
                invoice = invoice_cache.get(invoice_key) or PurchaseInvoice.query.filter_by(invoice_number=row["invoice_number"], series=row["series"], supplier_id=supplier.id if supplier else None).first()
                if invoice is None:
                    invoice = PurchaseInvoice(invoice_number=row["invoice_number"], series=row["series"], supplier_id=supplier.id if supplier else None, supplier_raw=row["supplier_raw"], invoice_date=row["invoice_date"], invoice_value=row["line_value"], status="RECEBIDA", imported=True, import_batch_id=batch.id)
                    db.session.add(invoice)
                    db.session.flush()
                    created["invoices"] += 1
                else:
                    invoice.invoice_value = (invoice.invoice_value or Decimal("0")) + row["line_value"]
                invoice_cache[invoice_key] = invoice
                if order is not None:
                    link = InvoicePurchaseOrderLink.query.filter_by(invoice_id=invoice.id, purchase_order_id=order.id).first()
                    if link is None:
                        db.session.add(InvoicePurchaseOrderLink(invoice_id=invoice.id, purchase_order_id=order.id, linked_value=row["line_value"]))
                        created["invoice_links"] += 1
                    if order_item is not None and not PurchaseInvoiceItem.query.filter_by(invoice_id=invoice.id, purchase_order_item_id=order_item.id).first():
                        db.session.add(PurchaseInvoiceItem(invoice_id=invoice.id, purchase_order_item_id=order_item.id, quantity_invoiced=row["quantity_ordered"] or row["quantity_requested"], quantity_received=row["quantity_received"], accepted=True))

            source_hash = hashlib.sha256(json.dumps(row["payload"], ensure_ascii=False, sort_keys=True).encode("utf-8")).hexdigest()
            source_row = PurchaseImportSourceRow(batch_id=batch.id, source_row_number=row["source_row_number"], source_hash=source_hash, source_payload=row["payload"], normalized_entity_ids={"purchase_request_id": request.id, "purchase_request_item_id": item.id, "purchase_order_id": order.id if order else None, "purchase_invoice_id": invoice.id if row["invoice_number"] else None})
            db.session.add(source_row)

        request.request_type = "MISTO" if material_count and service_count else ("MATERIAL" if material_count else "SERVICO")
        requested_total = max(1, int(sum((row["quantity_requested"] for row in request_rows), Decimal("0"))))
        received_total = int(sum((row["quantity_received"] for row in request_rows), Decimal("0")))
        request.requested_quantity = requested_total
        request.received_quantity = min(max(0, received_total), requested_total)
        if received_total > requested_total:
            request.data_quality_flags = sorted(set((request.data_quality_flags or []) + ["RECEBIDO_ACIMA_DO_SOLICITADO"]))
        request.material_id = next((item.material_id for item in request.items if item.material_id), None)
        db.session.add(PurchaseProcessEvent(entity_type="PURCHASE_REQUEST", entity_id=request.id, event_type="IMPORTADO", new_status=request.status, actor_id=user_id, comment="Importação histórica da fonte de compras."))

    batch.rows_created = sum(created.values())
    batch.status = "CONCLUIDO"
    batch.finished_at = now_manaus_naive()
    db.session.commit()
    return {"status": "CONCLUIDO", "batch": batch.to_dict(), "created": created, "reconciliation": {"source_rows": len(rows), "purchase_requests": len(grouped), "materials": created["materials"], "services": created["services"], "purchase_orders": created["orders"], "purchase_invoices": created["invoices"]}}
